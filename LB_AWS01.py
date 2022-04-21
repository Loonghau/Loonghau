# -*- coding: utf-8 -*-
"""
@Time : 2020/12/29 15:51
@Author : Loonghau XU
@FileName: tem_hum_new.py
@SoftWare: PyCharm
"""
import os
import sys
import time

import serial
import serial.tools.list_ports
from PyQt5 import QtWidgets
from PyQt5.Qt import QThread, pyqtSignal, QMutex
from PyQt5.QtGui import QDoubleValidator
from openpyxl import load_workbook, Workbook

from motion import Action
from temperature import Ui_Form
from zmcwrapper import ZMCWrapper

global tem, hum, ban, ban1

qmut_1 = QMutex()  # 创建线程锁
qmut_2 = QMutex()
qmut_3 = QMutex()  # 创建线程锁
qmut_4 = QMutex()

ac = Action()
zmc = ZMCWrapper()  # 实例化类
ip_2048 = "192.168.0.11"
zmc.connect(ip_2048)  # 连接控制器
# zmc.read_info()  # 读取控制器信息
zmc.set_axis_type()
zmc.set_axis_units()


class Pyqt5_Serial(QtWidgets.QWidget, Ui_Form):
    def __init__(self):
        super(Pyqt5_Serial, self).__init__()
        self.setupUi(self)
        self.setWindowTitle("全自动恒温恒湿称重系统")
        self.mythread = MyThread()  # 实例化线程
        self.mythread.signal_T.connect(
            self.callback_T)  # 连接线程类中自定义信号槽到本类的自定义槽函数
        self.mythread.signal_H.connect(self.callback_H)
        self.mythread.signal_B.connect(self.callback_B)
        self.mythread.start()  # 开启线程不是调用run函数而是调用start函数
        self.slot_init()  # 设置槽函数

    def callback_T(self, tem):
        doubleValidator = QDoubleValidator(0, 50, 1)
        self.lineEdit_2.setValidator(doubleValidator)
        self.lineEdit_2.setText(tem)

    def callback_H(self, hum):
        doubleValidator = QDoubleValidator(0, 50, 1)
        self.lineEdit_3.setValidator(doubleValidator)
        self.lineEdit_3.setText(hum)

    def callback_B(self, ban):
        self.lineEdit.setText(ban)

    def callback_I(self, csnum):
        self.lineEdit_4.setText(csnum)

    def callback_T0(self, now_time_0):
        self.lineEdit_5.setText(now_time_0)

    def callback_T1(self, now_time_1):
        self.lineEdit_6.setText(now_time_1)

    def slot_init(self):  # 设置槽函数
        self.pushButton.clicked.connect(self.weight)
        self.pushButton_4.clicked.connect(self.weight2)
        self.pushButton_2.clicked.connect(self.stop)
        self.pushButton_3.clicked.connect(self.begin)
        self.pushButton_5.clicked.connect(self.begin2)
        self.pushButton_6.clicked.connect(self.stop_2)
        self.pushButton_7.clicked.connect(self.code)
        self.pushButton_8.clicked.connect(self.ion_fan)

    def weight(self):
        self.thread_1 = Thread_1()  # 创建线程
        self.thread_1.signal_I.connect(self.callback_I)
        self.thread_1.signal_T0.connect(self.callback_T0)
        self.thread_1.signal_T1.connect(self.callback_T1)
        self.thread_1.start()  # 开始线程

    def weight2(self):
        self.thread_2 = Thread_2()  # 创建线程
        self.thread_2.signal_I.connect(self.callback_I)
        self.thread_2.signal_T0.connect(self.callback_T0)
        self.thread_2.signal_T1.connect(self.callback_T1)
        self.thread_2.start()  # 开始线程

    def stop(self):
        self.thread_3 = Thread_3()  # 创建线程
        self.thread_3.start()  # 开始线程

    def begin(self):
        self.thread_4 = Thread_4()  # 创建线程
        self.thread_4.start()  # 开始线程

    def begin2(self):
        self.thread_2_B = Thread_5()
        self.thread_2_B.start()

    def stop_2(self):
        self.thread_2_S = Thread_6()
        self.thread_2_S.start()

    def code(self):
        self.thread_2_C = Thread_7()
        self.thread_2_C.start()

    def ion_fan(self):
        self.thread_2_ION = Thread_8()
        self.thread_2_ION.start()


class Thread_1(QThread):
    global weight
    signal_I = pyqtSignal(str)
    signal_T0 = pyqtSignal(str)
    signal_T1 = pyqtSignal(str)

    def __init__(self):
        super().__init__()

    def run(self):
        # qmut_1.lock()
        ip_2048 = "192.168.0.11"
        zmc.connect(ip_2048)  # 连接控制器
        # zmc.read_info()  # 读取控制器信息
        zmc.set_axis_type()
        zmc.set_axis_units()
        now_time_0 = time.strftime(
            '%Y-%m-%d %H-%M',
            time.localtime(
                time.time()))
        self.signal_T0.emit(str(now_time_0))
        # self.signal_T1.emit( )
        zmc.set_op(7, 0)
        zmc.axis_1_up(25000)  # Z轴升高至第一层样品下位置
        time.sleep(8)  # 时间ok
        zmc.axis_0_move(-2642)  # Z轴旋转至第一层样品下位置
        time.sleep(5)  # 时间ok
        for i in range(1, 15):
            print(i)
            if zmc.get_input(10) == 0:
                time.sleep(2)
                zmc.axis_2_move(8572)
                time.sleep(5)
                continue
            else:
                csnum = str(chr(8544)) + "-" + str(i)
                self.signal_I.emit(csnum)
                zmc.axis_1_up(43000)  # Z轴升高，托住样品
                if zmc.get_op(7) == 1:
                    break
                time.sleep(6)  # 时间ok
                zmc.axis_0_move(0)  # Z轴旋转至初始位置
                if zmc.get_op(7) == 1:
                    break
                time.sleep(4)
                zmc.axis_1_up(50000)
                if zmc.get_op(7) == 1:
                    break
                time.sleep(3)
                zmc.set_op(1, 1)  # 离子风开
                time.sleep(5)  # 离子风吹5S时间
                zmc.set_op(1, 0)  # 离子风关

                y = serial.Serial('com30', 115200, timeout=1)  # 扫码
                myinput = bytes([0X16, 0X54, 0X0D])  # 需要发送"S"
                y.write(myinput)  # 用write函数向串口发送数据
                try:
                    myout = y.read(13)  # 提取接收缓冲区中的前7个字节数
                    ban = int(myout)
                    # print(ban)
                    y.close()
                except BaseException:
                    ban = " "
                    # print(ban)
                    y.close()

                time.sleep(2)
                zmc.axis_0_move(7627)
                if zmc.get_op(7) == 1:
                    break
                time.sleep(9)
                zmc.axis_1_up(35000)
                if zmc.get_op(7) == 1:
                    break
                time.sleep(6)
                zmc.axis_0_move(11000)
                if zmc.get_op(7) == 1:
                    break
                time.sleep(5)
                zmc.axis_3_down()
                if zmc.get_op(7) == 1:
                    break
                time.sleep(60)
                x = serial.Serial('com21', 2400, timeout=1)  # 温湿度
                z = serial.Serial('com29', 9600, timeout=1)  # 天平
                # 需要发送的十六进制数据
                myinput = bytes([0X53, 0X30, 0X31, 0X39, 0X39, 0X41])
                x.write(myinput)  # 用write函数向串口发送数据
                myinput1 = bytes([0X53, 0X0A, 0X0D])  # 需要发送"S"
                z.write(myinput1)  # 用write函数向串口发送数据
                myout = x.read(25)  # 提取接收缓冲区中的前7个字节数
                datas = ''.join(
                    map(lambda x: ('/x' if len(hex(x)) >= 4 else '/x0') + hex(x)[2:], myout))
                new_datas = datas[2:].split('/x')  # 由于datas变量中的数据前两

                tem = (int(new_datas[3],
                           16) + int(new_datas[4],
                                     16) * 256 + int(new_datas[5],
                                                     16) * 65536 + int(new_datas[6],
                       16) * 16777216) / 100
                hum = (int(new_datas[11],
                           16) + int(new_datas[12],
                                     16) * 256 + int(new_datas[13],
                                                     16) * 65536 + int(new_datas[14],
                       16) * 16777216) / 100
                myout1 = z.read(25)  # 提取接收缓冲区中的前7个字节数
                weight = str(myout1[6:14], encoding='utf-8')
                print(weight + "g")
                x.close()
                z.close()
                time.sleep(1)
                excelName = 'text.xlsx'

                if os.path.exists(excelName):
                    wb = load_workbook(excelName)
                    pass
                else:
                    wb = Workbook()

                now_time = time.strftime('%m-%d', time.localtime(time.time()))
                if wb.sheetnames[0] == str(now_time):
                    ws2 = wb[str(now_time)]
                    pass
                else:
                    ws2 = wb.create_sheet(str(now_time), 0)
                    ws2['A1'] = '日期/时间'
                    ws2['B1'] = '样品编号'
                    ws2['C1'] = '温度'
                    ws2['D1'] = '湿度'
                    ws2['E1'] = '重量'
                    ws2['F1'] = '样品位号'
                now_time_1 = time.strftime(
                    '%Y-%m-%d %H-%M',
                    time.localtime(
                        time.time()))
                ws2.column_dimensions['A'].width = 20
                ws2.column_dimensions['B'].width = 15
                ws2.column_dimensions['C'].width = 15
                ws2.column_dimensions['D'].width = 15
                ws2.column_dimensions['E'].width = 15
                ws2.column_dimensions['F'].width = 15
                list = [str(now_time_1), ban, tem, hum, weight + "g", csnum]
                # print(list)
                ws2.append(list)
                wb.save(excelName)

                zmc.axis_3_up()
                if zmc.get_op(7) == 1:
                    break
                time.sleep(25)
                zmc.axis_0_move(7627)
                if zmc.get_op(7) == 1:
                    break
                time.sleep(5)
                zmc.axis_1_up(50000)
                if zmc.get_op(7) == 1:
                    break
                time.sleep(6)
                zmc.axis_0_move(0)
                if zmc.get_op(7) == 1:
                    break
                time.sleep(9)
                zmc.axis_1_up(43000)
                if zmc.get_op(7) == 1:
                    break
                time.sleep(5)
                zmc.axis_0_move(-2642)
                if zmc.get_op(7) == 1:
                    break
                time.sleep(5)
                zmc.axis_1_up(25000)
                if zmc.get_op(7) == 1:
                    break
                time.sleep(5)
                zmc.axis_2_move(8572)
                if zmc.get_op(7) == 1:
                    break
                time.sleep(4)
        zmc.axis_0_move(0)
        time.sleep(4)
        zmc.axis_1_up(134000)
        time.sleep(38)
        zmc.axis_0_move(-2642)
        time.sleep(5)
        for i in range(1, 15):
            print(i)
            if zmc.get_input(10) == 0:
                time.sleep(2)
                zmc.axis_2_move(8572)
                time.sleep(5)
                continue
            else:
                csnum = str(chr(8545)) + "-" + str(i)
                self.signal_I.emit(csnum)
                zmc.axis_1_up(150000)  # Z轴升高，托住样品
                if zmc.get_op(7) == 1:
                    break
                time.sleep(6)
                zmc.axis_0_move(0)  # Z轴旋转至初始位置
                if zmc.get_op(7) == 1:
                    break
                time.sleep(5)
                zmc.axis_1_up(50000)
                if zmc.get_op(7) == 1:
                    break
                time.sleep(30)
                zmc.set_op(1, 1)
                time.sleep(5)
                zmc.set_op(1, 0)

                y = serial.Serial('com30', 115200, timeout=1)  # 扫码
                myinput = bytes([0X16, 0X54, 0X0D])  # 需要发送"S"
                y.write(myinput)  # 用write函数向串口发送数据
                try:
                    myout = y.read(13)  # 提取接收缓冲区中的前7个字节数
                    ban = int(myout)
                    print(ban)
                    y.close()
                except BaseException:
                    ban = " "
                    print(ban)
                    y.close()

                time.sleep(2)
                zmc.axis_0_move(7627)
                if zmc.get_op(7) == 1:
                    break
                time.sleep(9)
                zmc.axis_1_up(35000)
                if zmc.get_op(7) == 1:
                    break
                time.sleep(6)
                zmc.axis_0_move(11000)
                if zmc.get_op(7) == 1:
                    break
                time.sleep(5)
                zmc.axis_3_down()
                if zmc.get_op(7) == 1:
                    break
                time.sleep(60)
                x = serial.Serial('com21', 2400, timeout=1)
                z = serial.Serial('com29', 9600, timeout=1)
                # 需要发送的十六进制数据
                myinput = bytes([0X53, 0X30, 0X31, 0X39, 0X39, 0X41])
                x.write(myinput)  # 用write函数向串口发送数据
                myinput1 = bytes([0X53, 0X0A, 0X0D])  # 需要发送"S"
                z.write(myinput1)  # 用write函数向串口发送数据
                myout = x.read(25)  # 提取接收缓冲区中的前7个字节数
                datas = ''.join(
                    map(lambda x: ('/x' if len(hex(x)) >= 4 else '/x0') + hex(x)[2:], myout))
                new_datas = datas[2:].split('/x')  # 由于datas变量中的数据前两

                tem = (int(new_datas[3],
                           16) + int(new_datas[4],
                                     16) * 256 + int(new_datas[5],
                                                     16) * 65536 + int(new_datas[6],
                       16) * 16777216) / 100
                hum = (int(new_datas[11],
                           16) + int(new_datas[12],
                                     16) * 256 + int(new_datas[13],
                                                     16) * 65536 + int(new_datas[14],
                       16) * 16777216) / 100
                myout1 = z.read(25)  # 提取接收缓冲区中的前7个字节数
                weight = str(myout1[6:14], encoding='utf-8')
                print(weight)
                x.close()
                z.close()
                time.sleep(1)
                excelName = 'text.xlsx'

                if os.path.exists(excelName):
                    wb = load_workbook(excelName)
                    pass
                else:
                    wb = Workbook()

                now_time = time.strftime('%m-%d', time.localtime(time.time()))
                if wb.sheetnames[0] == str(now_time):
                    ws2 = wb[str(now_time)]
                    pass
                else:
                    ws2 = wb.create_sheet(str(now_time), 0)
                    ws2['A1'] = '日期/时间'
                    ws2['B1'] = '样品编号'
                    ws2['C1'] = '温度'
                    ws2['D1'] = '湿度'
                    ws2['E1'] = '重量'
                    ws2['F1'] = '样品位号'
                now_time_1 = time.strftime(
                    '%Y-%m-%d %H-%M',
                    time.localtime(
                        time.time()))
                ws2.column_dimensions['A'].width = 20
                ws2.column_dimensions['B'].width = 15
                ws2.column_dimensions['C'].width = 15
                ws2.column_dimensions['D'].width = 15
                ws2.column_dimensions['E'].width = 15
                ws2.column_dimensions['F'].width = 15
                list = [str(now_time_1), ban, tem, hum, weight + "g", csnum]
                # print(list)
                ws2.append(list)
                wb.save(excelName)

                zmc.axis_3_up()
                if zmc.get_op(7) == 1:
                    break
                time.sleep(25)
                zmc.axis_0_move(7627)
                if zmc.get_op(7) == 1:
                    break
                time.sleep(5)
                zmc.axis_1_up(50000)
                if zmc.get_op(7) == 1:
                    break
                time.sleep(6)
                zmc.axis_0_move(0)
                if zmc.get_op(7) == 1:
                    break
                time.sleep(9)
                zmc.axis_1_up(150000)
                if zmc.get_op(7) == 1:
                    break
                time.sleep(30)
                zmc.axis_0_move(-2642)
                if zmc.get_op(7) == 1:
                    break
                time.sleep(4)
                zmc.axis_1_up(134000)
                if zmc.get_op(7) == 1:
                    break
                time.sleep(5)
                zmc.axis_2_move(8572)
                if zmc.get_op(7) == 1:
                    break
                time.sleep(4)
        zmc.axis_0_move(0)
        time.sleep(5)
        zmc.axis_1_back()
        now_time_1 = time.strftime(
            '%Y-%m-%d %H-%M',
            time.localtime(
                time.time()))
        self.signal_T1.emit(str(now_time_1))
        # qmut_1.unlock()


class Thread_2(QThread):
    global weight
    signal_I = pyqtSignal(str)
    signal_T0 = pyqtSignal(str)
    signal_T1 = pyqtSignal(str)

    def __init__(self):
        super().__init__()

    def run(self):
        # qmut_2.lock()
        ip_2048 = "192.168.0.11"
        zmc.connect(ip_2048)  # 连接控制器
        # zmc.read_info()  # 读取控制器信息
        zmc.set_axis_type()
        zmc.set_axis_units()
        time.sleep(7200)
        now_time_0 = time.strftime(
            '%Y-%m-%d %H-%M',
            time.localtime(
                time.time()))
        self.signal_T0.emit(str(now_time_0))
        # self.signal_T1.emit( )
        zmc.set_op(7, 0)
        zmc.axis_1_up(25000)  # Z轴升高至第一层样品下位置
        time.sleep(8)  # 时间ok
        zmc.axis_0_move(-2642)  # Z轴旋转至第一层样品下位置
        time.sleep(5)  # 时间ok
        for i in range(1, 15):
            print(i)
            if zmc.get_input(10) == 0:
                time.sleep(2)
                zmc.axis_2_move(8572)
                time.sleep(5)
                continue
            else:
                csnum = str(chr(8544)) + "-" + str(i)
                self.signal_I.emit(csnum)
                zmc.axis_1_up(43000)  # Z轴升高，托住样品
                if zmc.get_op(7) == 1:
                    break
                time.sleep(6)  # 时间ok
                zmc.axis_0_move(0)  # Z轴旋转至初始位置
                if zmc.get_op(7) == 1:
                    break
                time.sleep(4)
                zmc.axis_1_up(50000)
                if zmc.get_op(7) == 1:
                    break
                time.sleep(3)
                zmc.set_op(1, 1)  # 离子风开
                time.sleep(5)  # 离子风吹5S时间
                zmc.set_op(1, 0)  # 离子风关

                y = serial.Serial('com30', 115200, timeout=1)  # 扫码
                myinput = bytes([0X16, 0X54, 0X0D])  # 需要发送"S"
                y.write(myinput)  # 用write函数向串口发送数据
                try:
                    myout = y.read(13)  # 提取接收缓冲区中的前7个字节数
                    ban = int(myout)
                    # print(ban)
                    y.close()
                except BaseException:
                    ban = " "
                    # print(ban)
                    y.close()

                time.sleep(2)
                zmc.axis_0_move(7627)
                if zmc.get_op(7) == 1:
                    break
                time.sleep(9)
                zmc.axis_1_up(35000)
                if zmc.get_op(7) == 1:
                    break
                time.sleep(6)
                zmc.axis_0_move(11000)
                if zmc.get_op(7) == 1:
                    break
                time.sleep(5)
                zmc.axis_3_down()
                if zmc.get_op(7) == 1:
                    break
                time.sleep(60)
                x = serial.Serial('com21', 2400, timeout=1)  # 温湿度
                z = serial.Serial('com29', 9600, timeout=1)  # 天平
                # 需要发送的十六进制数据
                myinput = bytes([0X53, 0X30, 0X31, 0X39, 0X39, 0X41])
                x.write(myinput)  # 用write函数向串口发送数据
                myinput1 = bytes([0X53, 0X0A, 0X0D])  # 需要发送"S"
                z.write(myinput1)  # 用write函数向串口发送数据
                myout = x.read(25)  # 提取接收缓冲区中的前7个字节数
                datas = ''.join(
                    map(lambda x: ('/x' if len(hex(x)) >= 4 else '/x0') + hex(x)[2:], myout))
                new_datas = datas[2:].split('/x')  # 由于datas变量中的数据前两

                tem = (int(new_datas[3],
                           16) + int(new_datas[4],
                                     16) * 256 + int(new_datas[5],
                                                     16) * 65536 + int(new_datas[6],
                       16) * 16777216) / 100
                hum = (int(new_datas[11],
                           16) + int(new_datas[12],
                                     16) * 256 + int(new_datas[13],
                                                     16) * 65536 + int(new_datas[14],
                       16) * 16777216) / 100
                myout1 = z.read(25)  # 提取接收缓冲区中的前7个字节数
                weight = str(myout1[6:14], encoding='utf-8')
                print(weight + "g")
                x.close()
                z.close()
                time.sleep(1)
                excelName = 'text.xlsx'

                if os.path.exists(excelName):
                    wb = load_workbook(excelName)
                    pass
                else:
                    wb = Workbook()

                now_time = time.strftime('%m-%d', time.localtime(time.time()))
                if wb.sheetnames[0] == str(now_time):
                    ws2 = wb[str(now_time)]
                    pass
                else:
                    ws2 = wb.create_sheet(str(now_time), 0)
                    ws2['A1'] = '日期/时间'
                    ws2['B1'] = '样品编号'
                    ws2['C1'] = '温度'
                    ws2['D1'] = '湿度'
                    ws2['E1'] = '重量'
                    ws2['F1'] = '样品位号'
                now_time_1 = time.strftime(
                    '%Y-%m-%d %H-%M',
                    time.localtime(
                        time.time()))
                ws2.column_dimensions['A'].width = 20
                ws2.column_dimensions['B'].width = 15
                ws2.column_dimensions['C'].width = 15
                ws2.column_dimensions['D'].width = 15
                ws2.column_dimensions['E'].width = 15
                ws2.column_dimensions['F'].width = 15
                list = [str(now_time_1), ban, tem, hum, weight + "g", csnum]
                # print(list)
                ws2.append(list)
                wb.save(excelName)

                zmc.axis_3_up()
                if zmc.get_op(7) == 1:
                    break
                time.sleep(25)
                zmc.axis_0_move(7627)
                if zmc.get_op(7) == 1:
                    break
                time.sleep(5)
                zmc.axis_1_up(50000)
                if zmc.get_op(7) == 1:
                    break
                time.sleep(6)
                zmc.axis_0_move(0)
                if zmc.get_op(7) == 1:
                    break
                time.sleep(9)
                zmc.axis_1_up(43000)
                if zmc.get_op(7) == 1:
                    break
                time.sleep(5)
                zmc.axis_0_move(-2642)
                if zmc.get_op(7) == 1:
                    break
                time.sleep(5)
                zmc.axis_1_up(25000)
                if zmc.get_op(7) == 1:
                    break
                time.sleep(5)
                zmc.axis_2_move(8572)
                if zmc.get_op(7) == 1:
                    break
                time.sleep(4)
        zmc.axis_0_move(0)
        time.sleep(4)
        zmc.axis_1_up(134000)
        time.sleep(38)
        zmc.axis_0_move(-2642)
        time.sleep(5)
        for i in range(1, 15):
            print(i)
            if zmc.get_input(10) == 0:
                time.sleep(2)
                zmc.axis_2_move(8572)
                time.sleep(5)
                continue
            else:
                csnum = str(chr(8545)) + "-" + str(i)
                self.signal_I.emit(csnum)
                zmc.axis_1_up(150000)  # Z轴升高，托住样品
                if zmc.get_op(7) == 1:
                    break
                time.sleep(6)
                zmc.axis_0_move(0)  # Z轴旋转至初始位置
                if zmc.get_op(7) == 1:
                    break
                time.sleep(5)
                zmc.axis_1_up(50000)
                if zmc.get_op(7) == 1:
                    break
                time.sleep(30)
                zmc.set_op(1, 1)
                time.sleep(5)
                zmc.set_op(1, 0)

                y = serial.Serial('com30', 115200, timeout=1)  # 扫码
                myinput = bytes([0X16, 0X54, 0X0D])  # 需要发送"S"
                y.write(myinput)  # 用write函数向串口发送数据
                try:
                    myout = y.read(13)  # 提取接收缓冲区中的前7个字节数
                    ban = int(myout)
                    print(ban)
                    y.close()
                except BaseException:
                    ban = " "
                    print(ban)
                    y.close()

                time.sleep(2)
                zmc.axis_0_move(7627)
                if zmc.get_op(7) == 1:
                    break
                time.sleep(9)
                zmc.axis_1_up(35000)
                if zmc.get_op(7) == 1:
                    break
                time.sleep(6)
                zmc.axis_0_move(11000)
                if zmc.get_op(7) == 1:
                    break
                time.sleep(5)
                zmc.axis_3_down()
                if zmc.get_op(7) == 1:
                    break
                time.sleep(60)
                x = serial.Serial('com21', 2400, timeout=1)
                z = serial.Serial('com29', 9600, timeout=1)
                # 需要发送的十六进制数据
                myinput = bytes([0X53, 0X30, 0X31, 0X39, 0X39, 0X41])
                x.write(myinput)  # 用write函数向串口发送数据
                myinput1 = bytes([0X53, 0X0A, 0X0D])  # 需要发送"S"
                z.write(myinput1)  # 用write函数向串口发送数据
                myout = x.read(25)  # 提取接收缓冲区中的前7个字节数
                datas = ''.join(
                    map(lambda x: ('/x' if len(hex(x)) >= 4 else '/x0') + hex(x)[2:], myout))
                new_datas = datas[2:].split('/x')  # 由于datas变量中的数据前两

                tem = (int(new_datas[3],
                           16) + int(new_datas[4],
                                     16) * 256 + int(new_datas[5],
                                                     16) * 65536 + int(new_datas[6],
                       16) * 16777216) / 100
                hum = (int(new_datas[11],
                           16) + int(new_datas[12],
                                     16) * 256 + int(new_datas[13],
                                                     16) * 65536 + int(new_datas[14],
                       16) * 16777216) / 100
                myout1 = z.read(25)  # 提取接收缓冲区中的前7个字节数
                weight = str(myout1[6:14], encoding='utf-8')
                print(weight)
                x.close()
                z.close()
                time.sleep(1)
                excelName = 'text.xlsx'

                if os.path.exists(excelName):
                    wb = load_workbook(excelName)
                    pass
                else:
                    wb = Workbook()

                now_time = time.strftime('%m-%d', time.localtime(time.time()))
                if wb.sheetnames[0] == str(now_time):
                    ws2 = wb[str(now_time)]
                    pass
                else:
                    ws2 = wb.create_sheet(str(now_time), 0)
                    ws2['A1'] = '日期/时间'
                    ws2['B1'] = '样品编号'
                    ws2['C1'] = '温度'
                    ws2['D1'] = '湿度'
                    ws2['E1'] = '重量'
                    ws2['F1'] = '样品位号'
                now_time_1 = time.strftime(
                    '%Y-%m-%d %H-%M',
                    time.localtime(
                        time.time()))
                ws2.column_dimensions['A'].width = 20
                ws2.column_dimensions['B'].width = 15
                ws2.column_dimensions['C'].width = 15
                ws2.column_dimensions['D'].width = 15
                ws2.column_dimensions['E'].width = 15
                ws2.column_dimensions['F'].width = 15
                list = [str(now_time_1), ban, tem, hum, weight + "g", csnum]
                # print(list)
                ws2.append(list)
                wb.save(excelName)

                zmc.axis_3_up()
                if zmc.get_op(7) == 1:
                    break
                time.sleep(25)
                zmc.axis_0_move(7627)
                if zmc.get_op(7) == 1:
                    break
                time.sleep(5)
                zmc.axis_1_up(50000)
                if zmc.get_op(7) == 1:
                    break
                time.sleep(6)
                zmc.axis_0_move(0)
                if zmc.get_op(7) == 1:
                    break
                time.sleep(9)
                zmc.axis_1_up(150000)
                if zmc.get_op(7) == 1:
                    break
                time.sleep(30)
                zmc.axis_0_move(-2642)
                if zmc.get_op(7) == 1:
                    break
                time.sleep(4)
                zmc.axis_1_up(134000)
                if zmc.get_op(7) == 1:
                    break
                time.sleep(5)
                zmc.axis_2_move(8572)
                if zmc.get_op(7) == 1:
                    break
                time.sleep(4)
        zmc.axis_0_move(0)
        time.sleep(5)
        zmc.axis_1_back()
        now_time_1 = time.strftime(
            '%Y-%m-%d %H-%M',
            time.localtime(
                time.time()))
        self.signal_T1.emit(str(now_time_1))
        # qmut_2.unlock()


class Thread_3(QThread):
    def __init__(self):
        super().__init__()

    def run(self):
        # qmut_3.lock()
        ip_2048 = "192.168.0.11"
        zmc.connect(ip_2048)  # 连接控制器
        # zmc.read_info()  # 读取控制器信息
        zmc.set_axis_type()
        zmc.set_axis_units()
        zmc.set_op(7, 1)
        time.sleep(1000)
        # zmc.set_op(7, 0)
        # qmut_3.unlock()


class Thread_4(QThread):  # 初始化状态，归零
    def __init__(self):
        super().__init__()

    def run(self):
        # qmut_4.lock()
        ip_2048 = "192.168.0.11"
        zmc.connect(ip_2048)  # 连接控制器
        # zmc.read_info()  # 读取控制器信息
        zmc.set_axis_type()
        zmc.set_axis_units()
        zmc.set_op(7, 0)
        zmc.axis_2_back()
        time.sleep(zmc.get_DPOS(2) / (zmc.get_speed(2)))
        # zmc.axis_2_move(1600)
        if zmc.get_input(8) == 0:
            zmc.axis_3_up()
            time.sleep(20)
            if zmc.get_DPOS(0) > 1200:
                zmc.axis_1_up(51000)
                time.sleep(10)
                zmc.set_op(0, 1)
                time.sleep(8)
                zmc.set_op(0, 0)
                zmc.axis_1_back()
                time.sleep(5)
            else:
                zmc.axis_0_move(0)
                time.sleep(5)
                zmc.axis_1_back()
        else:
            if zmc.get_DPOS(0) > 1200:
                zmc.axis_1_up(51000)
                time.sleep(10)
                zmc.set_op(0, 1)
                time.sleep(8)
                zmc.set_op(0, 0)
                zmc.axis_1_back()
                time.sleep(5)
            else:
                zmc.axis_0_move(0)
                time.sleep(5)
                zmc.axis_1_back()
        zmc.set_DPOS(0, 0)
        zmc.set_DPOS(1, 0)
        # qmut_4.unlock()


class Thread_5(QThread):
    def __init__(self):
        super().__init__()

    def run(self):
        ip_2048 = "192.168.0.11"
        zmc.connect(ip_2048)  # 连接控制器
        # zmc.read_info()  # 读取控制器信息
        zmc.set_axis_type()
        zmc.set_axis_units()
        zmc.set_op(7, 0)
        zmc.set_move(2, 1)


class Thread_6(QThread):
    def __init__(self):
        super().__init__()

    def run(self):
        ip_2048 = "192.168.0.11"
        zmc.connect(ip_2048)  # 连接控制器
        # zmc.read_info()  # 读取控制器信息
        zmc.set_axis_type()
        zmc.set_axis_units()
        zmc.set_op(7, 0)
        zmc.set_cancel(2, 0)


class Thread_7(QThread):
    def __init__(self):
        super().__init__()

    def run(self):
        y = serial.Serial('com30', 115200, timeout=1)
        myinput = bytes([0X16, 0X54, 0X0D])  # 需要发送"S"
        y.write(myinput)  # 用write函数向串口发送数据
        try:
            myout = y.read(13)  # 提取接收缓冲区中的前7个字节数
            ban = int(myout)
            print(ban)
            y.close()
        except BaseException:
            ban = " "
            print(ban)
            y.close()


class Thread_8(QThread):
    def __init__(self):
        super().__init__()

    def run(self):
        zmc.set_op(1, 1)  # 离子风开
        time.sleep(5)  # 离子风吹5S时间
        zmc.set_op(1, 0)  # 离子风关


class MyThread(QThread):
    # global tem, hum, ban
    signal_T = pyqtSignal(str)  # 自定义一个pyqtSignal信号,信号参数是个字符串str类型
    signal_H = pyqtSignal(str)
    signal_B = pyqtSignal(str)

    def __init__(self):
        super(MyThread, self).__init__()

    def run(self):
        for i in range(10000):
            x = serial.Serial('com21', 2400, timeout=1)
            # 需要发送的十六进制数据
            myinput = bytes([0X53, 0X30, 0X31, 0X39, 0X39, 0X41])
            x.write(myinput)  # 用write函数向串口发送数据
            myout = x.read(25)  # 提取接收缓冲区中的前7个字节数
            datas = ''.join(
                map(lambda x: ('/x' if len(hex(x)) >= 4 else '/x0') + hex(x)[2:], myout))
            new_datas = datas[2:].split('/x')  # 由于datas变量中的数据前两

            tem = (int(new_datas[3],
                       16) + int(new_datas[4],
                                 16) * 256 + int(new_datas[5],
                                                 16) * 65536 + int(new_datas[6],
                                                                   16) * 16777216) / 100
            hum = (int(new_datas[11],
                       16) + int(new_datas[12],
                                 16) * 256 + int(new_datas[13],
                                                 16) * 65536 + int(new_datas[14],
                                                                   16) * 16777216) / 100
            # print("现在温度：", tem)
            # print("现在湿度：", hum)
            self.signal_T.emit(str(tem))
            self.signal_H.emit(str(hum))

            y = serial.Serial('com29', 9600, timeout=1)
            myinput1 = bytes([0X53, 0X0A, 0X0D])  # 需要发送"S"
            y.write(myinput1)  # 用write函数向串口发送数据
            myout1 = y.read(25)  # 提取接收缓冲区中的前7个字节数
            ban = str(myout1[6:14], encoding='utf-8')
            # print("重量：", ban)
            self.signal_B.emit(str(ban))
            if zmc.get_input(9) == 1:
                x.close()
                y.close()
                time.sleep(65)
            else:
                x.close()
                y.close()
                time.sleep(1)


if __name__ == "__main__":
    app = QtWidgets.QApplication(sys.argv)
    myshow = Pyqt5_Serial()
    myshow.show()
    sys.exit(app.exec_())
