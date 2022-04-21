# -*- coding: utf-8 -*-
""" 
@Time : 2020/12/29 15:51
@Author : Loonghau XU
@FileName: LB_AWS02.py
@SoftWare: PyCharm
"""
import os
import sys
import time

import serial
import serial.tools.list_ports
import wmi
from PyQt5 import QtWidgets
from PyQt5.Qt import QThread, pyqtSignal, QMutex
from PyQt5.QtGui import QDoubleValidator
from openpyxl import load_workbook, Workbook

from temperature import Ui_Form
from zmcwrapper import ZMCWrapper

global tem, hum, ban, ban1

qmut_1 = QMutex()  # 创建线程锁
qmut_2 = QMutex()
qmut_3 = QMutex()  # 创建线程锁
qmut_4 = QMutex()

# ac = Action()
zmc = ZMCWrapper()  # 实例化类、封装控制器类函数
ip_2048 = "192.168.0.11"  # 控制器IP地址
zmc.connect(ip_2048)  # 连接控制器
mac = '00:9A:27:E0:6D:1D'  # 主机amc地址
# zmc.read_info()       # 读取控制器信息
zmc.set_axis_type()  # 设定控制器轴类型“7”
zmc.set_axis_units()  # 设定轴脉冲当量


def GetMac():
    global ret
    c = wmi.WMI()
    ret = c.Win32_NetworkAdapterConfiguration(IPEnabled=1)[0].MACAddress
    return ret


mac1 = GetMac()

if mac == mac1:
    class Pyqt5_Serial(QtWidgets.QWidget, Ui_Form):
        def __init__(self):
            super(Pyqt5_Serial, self).__init__()
            self.setupUi(self)
            self.setWindowTitle("全自动恒温恒湿称重系统V1.0")
            self.mythread = MyThread()  # 实例化线程
            self.mythread.signal_T.connect(self.callback_T)  # 连接线程类中自定义信号槽到本类的自定义槽函数
            self.mythread.signal_H.connect(self.callback_H)  #
            self.mythread.start()  # 开启线程不是调用run函数而是调用start函数
            self.slot_init()  # 设置槽函数

        def callback_T(self, tem):
            doubleValidator = QDoubleValidator(0, 50, 1)
            self.lineEdit_2.setValidator(doubleValidator)
            self.lineEdit_2.setText(tem)

        def callback_H(self, hum):
            doubleValidator = QDoubleValidator(0, 50, 1)
            self.lineEdit_3.setValidator(doubleValidator)
            self.lineEdit_3.setText(hum)

        def callback_W(self, weight):
            self.lineEdit.setText(weight)

        def callback_I(self, csnum):
            self.lineEdit_4.setText(csnum)

        def callback_T0(self, now_time_0):
            self.lineEdit_5.setText(now_time_0)

        def callback_T1(self, now_time_1):
            self.lineEdit_6.setText(now_time_1)

        def callback_M(self, ban):
            self.lineEdit_7.setText(ban)

        def slot_init(self):  # 设置槽函数
            self.pushButton.clicked.connect(self.weight)  # 手动开始称重
            self.pushButton_4.clicked.connect(self.weight2)  # 恒温恒湿后称重
            self.pushButton_2.clicked.connect(self.stop)  # 紧急停止称重
            self.pushButton_20.clicked.connect(self.begin)
            self.pushButton_5.clicked.connect(self.O_begin2)
            self.pushButton_6.clicked.connect(self.O_stop_2)
            self.pushButton_7.clicked.connect(self.code)
            self.pushButton_8.clicked.connect(self.ion_fan)
            self.pushButton_9.clicked.connect(self.ban_up)
            self.pushButton_10.clicked.connect(self.ban_down)
            self.pushButton_11.clicked.connect(self.z_left)
            self.pushButton_12.clicked.connect(self.z_right)
            self.pushButton_13.clicked.connect(self.z_up)
            self.pushButton_14.clicked.connect(self.z_down)
            self.pushButton_15.clicked.connect(self.z_lr_stop)
            self.pushButton_16.clicked.connect(self.z_ud_back)
            self.pushButton_17.clicked.connect(self.O_back)
            self.pushButton_18.clicked.connect(self.z_lr_back)
            self.pushButton_19.clicked.connect(self.z_ud_stop)

        def weight(self):
            self.thread_1 = Thread_1()  # 创建线程
            self.thread_1.signal_I.connect(self.callback_I)
            self.thread_1.signal_W.connect(self.callback_W)
            self.thread_1.signal_B.connect(self.callback_M)
            self.thread_1.signal_T0.connect(self.callback_T0)
            self.thread_1.signal_T1.connect(self.callback_T1)
            self.thread_1.start()  # 开始线程

        def weight2(self):
            self.thread_2 = Thread_2()  # 创建线程
            self.thread_2.signal_I.connect(self.callback_I)
            self.thread_2.signal_W.connect(self.callback_W)
            self.thread_2.signal_B.connect(self.callback_M)
            self.thread_2.signal_T0.connect(self.callback_T0)
            self.thread_2.signal_T1.connect(self.callback_T1)
            self.thread_2.start()  # 开始线程

        def code(self):
            self.thread_2_C = Thread_7()
            self.thread_2_C.signal_B.connect(self.callback_M)
            self.thread_2_C.start()

        def stop(self):
            self.thread_3 = Thread_3()  # 创建线程
            self.thread_3.start()  # 开始线程

        def begin(self):
            self.thread_4 = Thread_4()  # 创建线程
            self.thread_4.start()  # 开始线程

        def O_begin2(self):
            self.thread_2_B = Thread_5()
            self.thread_2_B.start()

        def O_stop_2(self):
            self.thread_2_S = Thread_6()
            self.thread_2_S.start()

        def ion_fan(self):
            self.thread_2_ION = Thread_8()
            self.thread_2_ION.start()

        def ban_up(self):
            self.thread_ban_up = Thread_9()
            self.thread_ban_up.start()

        def ban_down(self):
            self.thread_ban_down = Thread_10()
            self.thread_ban_down.start()

        def z_left(self):
            self.thread_z_left = Thread_11()
            self.thread_z_left.start()

        def z_right(self):
            self.thread_z_right = Thread_12()
            self.thread_z_right.start()

        def z_up(self):
            self.thread_z_up = Thread_13()
            self.thread_z_up.start()

        def z_down(self):
            self.thread_z_down = Thread_14()
            self.thread_z_down.start()

        def O_back(self):
            self.thread_O_back = Thread_15()
            self.thread_O_back.start()

        def z_lr_stop(self):
            self.thread_z_lr_stop = Thread_16()
            self.thread_z_lr_stop.start()

        def z_lr_back(self):
            self.thread_z_lr_back = Thread_17()
            self.thread_z_lr_back.start()

        def z_ud_back(self):
            self.thread_z_ud_back = Thread_18()
            self.thread_z_ud_back.start()

        def z_ud_stop(self):
            self.thread_z_ud_stop = Thread_19()
            self.thread_z_ud_stop.start()


    class Thread_1(QThread):
        global weight
        signal_I = pyqtSignal(str)
        signal_W = pyqtSignal(str)
        signal_B = pyqtSignal(str)
        signal_T0 = pyqtSignal(str)
        signal_T1 = pyqtSignal(str)

        def __init__(self):
            super().__init__()

        def run(self):
            # qmut_1.lock()
            ip_2048 = "192.168.0.11"
            zmc.connect(ip_2048)  # 连接控制器
            # zmc.read_info()  # 读取控制器信息
            zmc.set_axis_type()
            zmc.set_axis_units()
            self.signal_T1.emit('')
            now_time_0 = time.strftime('%Y-%m-%d %H-%M', time.localtime(time.time()))
            self.signal_T0.emit(str(now_time_0))
            zmc.set_op(7, 0)
            zmc.axis_1_up(25000)  # Z轴升高至第一层样品下位置
            time.sleep(8)  # 时间ok
            zmc.axis_0_move(-2642)  # Z轴旋转至第一层样品下位置
            time.sleep(5)  # 时间ok
            for i in range(1, 15):
                print(i)
                if zmc.get_input(10) == 0:
                    time.sleep(2)
                    self.signal_B.emit('')
                    self.signal_W.emit('')
                    self.signal_I.emit('')
                    zmc.axis_2_move(8572)
                    time.sleep(5)
                    continue
                else:
                    csnum = str(chr(8544)) + "-" + str(i)
                    self.signal_I.emit(csnum)
                    self.signal_W.emit('')
                    zmc.axis_1_up(43000)  # Z轴升高，托住样品
                    if zmc.get_input(15) == 1:  # 按下急停按钮
                        break
                    time.sleep(6)  # 时间ok
                    zmc.axis_0_move(0)  # Z轴旋转至初始位置
                    if zmc.get_input(15) == 1:
                        break
                    time.sleep(4)
                    zmc.axis_1_up(76000)
                    if zmc.get_input(15) == 1:
                        break
                    time.sleep(9)
                    zmc.set_op(1, 1)  # 离子风开
                    time.sleep(15)  # 离子风吹15S时间
                    zmc.set_op(1, 0)  # 离子风关

                    y = serial.Serial('com7', 115200, timeout=1)  # 扫码
                    myinput = bytes([0X16, 0X54, 0X0D])  # 需要发送"S"
                    y.write(myinput)  # 用write函数向串口发送数据
                    try:
                        myout = y.read(13)  # 提取接收缓冲区中的前7个字节数
                        ban = int(myout)
                        self.signal_B.emit(str(ban))
                        # print(ban)
                        y.close()
                    except:
                        ban = " "
                        # print(ban)
                        y.close()
                    time.sleep(2)
                    zmc.axis_0_move(8540)
                    if zmc.get_input(15) == 1:
                        break
                    time.sleep(10)
                    zmc.axis_1_up(62000)
                    if zmc.get_input(15) == 1:
                        break
                    time.sleep(6)
                    zmc.axis_0_move(12000)
                    if zmc.get_input(15) == 1:
                        break
                    time.sleep(5)
                    zmc.axis_3_down()
                    if zmc.get_input(15) == 1:
                        break
                    time.sleep(40)
                    th = serial.Serial('com6', 2400, timeout=1)  # 温湿度
                    z = serial.Serial('com2', 9600, timeout=1)  # 天平
                    myinput = bytes([0X53, 0X30, 0X31, 0X39, 0X39, 0X41])  # 需要发送的十六进制数据
                    th.write(myinput)  # 用write函数向串口发送数据
                    myinput1 = bytes([0X53, 0X0A, 0X0D])  # 需要发送"S"
                    z.write(myinput1)  # 用write函数向串口发送数据
                    myout = th.read(25)  # 提取接收缓冲区中的前7个字节数
                    datas = ''.join(map(lambda x: ('/x' if len(hex(x)) >= 4 else '/x0') + hex(x)[2:], myout))
                    new_datas = datas[2:].split('/x')  # 由于datas变量中的数据前两

                    tem = (int(new_datas[3], 16) + int(new_datas[4], 16) * 256 + int(new_datas[5], 16) * 65536 + int(
                        new_datas[6],
                        16) * 16777216) / 100
                    hum = (int(new_datas[11], 16) + int(new_datas[12], 16) * 256 + int(new_datas[13], 16) * 65536 + int(
                        new_datas[14],
                        16) * 16777216) / 100
                    myout1 = z.read(25)  # 提取接收缓冲区中的前7个字节数
                    weight = str(myout1[6:14], encoding='utf-8')
                    self.signal_W.emit(weight + 'g')
                    print(weight + "g")
                    th.close()
                    z.close()
                    time.sleep(1)
                    excelName1 = 'text.xlsx'
                    excelName = 'result.xlsx'

                    if os.path.exists(excelName1):
                        wb = load_workbook(excelName1)
                        pass
                    else:
                        wb = Workbook()

                    now_time = time.strftime('%m-%d', time.localtime(time.time()))
                    if wb.sheetnames[0] == str(now_time):
                        ws = wb[str(now_time)]
                        pass
                    else:
                        ws = wb.create_sheet(str(now_time), 0)
                        ws['A1'] = '日期/时间'
                        ws['B1'] = '样品编号'
                        ws['C1'] = '温度'
                        ws['D1'] = '湿度'
                        ws['E1'] = '重量'
                        ws['F1'] = '样品位号'
                    now_time_1 = time.strftime('%Y-%m-%d %H-%M', time.localtime(time.time()))
                    ws.column_dimensions['A'].width = 20
                    ws.column_dimensions['B'].width = 15
                    ws.column_dimensions['C'].width = 15
                    ws.column_dimensions['D'].width = 15
                    ws.column_dimensions['E'].width = 15
                    ws.column_dimensions['F'].width = 15
                    list = [str(now_time_1), str(ban), tem, hum, weight + "g", csnum]
                    # print(list)
                    ws.append(list)
                    wb.save(excelName1)
                    new_name = str(now_time_1) + str(csnum) + ".xlsx"
                    os.rename("text.xlsx", new_name)

                    if os.path.exists(excelName):
                        wb1 = load_workbook(excelName)
                        pass
                    else:
                        wb1 = Workbook()

                    now_time = time.strftime('%m-%d', time.localtime(time.time()))
                    if wb1.sheetnames[0] == str(now_time):
                        ws2 = wb1[str(now_time)]
                        pass
                    else:
                        ws2 = wb1.create_sheet(str(now_time), 0)
                        ws2['A1'] = '日期/时间'
                        ws2['B1'] = '样品编号'
                        ws2['C1'] = '温度'
                        ws2['D1'] = '湿度'
                        ws2['E1'] = '重量'
                        ws2['F1'] = '样品位号'
                    now_time_2 = time.strftime('%Y-%m-%d %H-%M', time.localtime(time.time()))
                    ws2.column_dimensions['A'].width = 20
                    ws2.column_dimensions['B'].width = 15
                    ws2.column_dimensions['C'].width = 15
                    ws2.column_dimensions['D'].width = 15
                    ws2.column_dimensions['E'].width = 15
                    ws2.column_dimensions['F'].width = 15
                    list1 = [str(now_time_2), str(ban), tem, hum, weight + "g", csnum]
                    # print(list)
                    ws2.append(list1)
                    wb1.save(excelName)

                    zmc.axis_3_up()
                    if zmc.get_input(15) == 1:
                        break
                    time.sleep(25)
                    zmc.axis_0_move(8540)
                    if zmc.get_input(15) == 1:
                        break
                    time.sleep(5)
                    zmc.axis_1_up(76000)
                    if zmc.get_input(15) == 1:
                        break
                    time.sleep(6)
                    zmc.axis_0_move(0)
                    if zmc.get_input(15) == 1:
                        break
                    time.sleep(9)
                    zmc.axis_1_up(43000)
                    if zmc.get_input(15) == 1:
                        break
                    time.sleep(9)
                    zmc.axis_0_move(-2642)
                    if zmc.get_input(15) == 1:
                        break
                    time.sleep(5)
                    zmc.axis_1_up(25000)
                    if zmc.get_input(15) == 1:
                        break
                    time.sleep(5)
                    zmc.axis_2_move(8572)
                    if zmc.get_input(15) == 1:
                        break
                    time.sleep(4)
            zmc.axis_0_move(0)
            time.sleep(4)
            zmc.axis_1_up(134000)
            time.sleep(38)
            zmc.axis_0_move(-2642)
            time.sleep(5)
            for i in range(1, 15):
                print(i)
                if zmc.get_input(10) == 0:
                    time.sleep(2)
                    self.signal_B.emit('')
                    self.signal_W.emit('')
                    self.signal_I.emit('')
                    zmc.axis_2_move(8572)
                    time.sleep(5)
                    continue
                else:
                    csnum = str(chr(8545)) + "-" + str(i)
                    self.signal_I.emit(csnum)
                    zmc.axis_1_up(150000)  # Z轴升高，托住样品
                    if zmc.get_input(15) == 1:
                        break
                    time.sleep(6)
                    zmc.axis_0_move(0)  # Z轴旋转至初始位置
                    if zmc.get_input(15) == 1:
                        break
                    time.sleep(5)
                    zmc.axis_1_up(76000)
                    if zmc.get_input(15) == 1:
                        break
                    time.sleep(20)
                    zmc.set_op(1, 1)
                    time.sleep(5)
                    zmc.set_op(1, 0)

                    y = serial.Serial('com7', 115200, timeout=1)  # 扫码
                    myinput = bytes([0X16, 0X54, 0X0D])  # 需要发送"S"
                    y.write(myinput)  # 用write函数向串口发送数据
                    try:
                        myout = y.read(13)  # 提取接收缓冲区中的前7个字节数
                        ban = int(myout)
                        self.signal_B.emit(str(ban))
                        print(ban)
                        y.close()
                    except:
                        ban = " "
                        print(ban)
                        y.close()

                    time.sleep(2)
                    zmc.axis_0_move(8540)
                    if zmc.get_input(15) == 1:
                        break
                    time.sleep(9)
                    zmc.axis_1_up(62000)
                    if zmc.get_input(15) == 1:
                        break
                    time.sleep(6)
                    zmc.axis_0_move(12000)
                    if zmc.get_input(15) == 1:
                        break
                    time.sleep(5)
                    zmc.axis_3_down()
                    if zmc.get_input(15) == 1:
                        break
                    time.sleep(60)
                    th = serial.Serial('com6', 2400, timeout=1)
                    z = serial.Serial('com2', 9600, timeout=1)
                    myinput = bytes([0X53, 0X30, 0X31, 0X39, 0X39, 0X41])  # 需要发送的十六进制数据
                    th.write(myinput)  # 用write函数向串口发送数据
                    myinput1 = bytes([0X53, 0X0A, 0X0D])  # 需要发送"S"
                    z.write(myinput1)  # 用write函数向串口发送数据
                    myout = th.read(25)  # 提取接收缓冲区中的前7个字节数
                    datas = ''.join(map(lambda x: ('/x' if len(hex(x)) >= 4 else '/x0') + hex(x)[2:], myout))
                    new_datas = datas[2:].split('/x')  # 由于datas变量中的数据前两

                    tem = (int(new_datas[3], 16) + int(new_datas[4], 16) * 256 + int(new_datas[5], 16) * 65536 + int(
                        new_datas[6],
                        16) * 16777216) / 100
                    hum = (int(new_datas[11], 16) + int(new_datas[12], 16) * 256 + int(new_datas[13], 16) * 65536 + int(
                        new_datas[14],
                        16) * 16777216) / 100
                    myout1 = z.read(25)  # 提取接收缓冲区中的前7个字节数
                    weight = str(myout1[6:14], encoding='utf-8')
                    self.signal_W.emit(weight + 'g')
                    print(weight)
                    th.close()
                    z.close()
                    time.sleep(1)
                    excelName1 = 'text.xlsx'
                    excelName = 'result.xlsx'

                    if os.path.exists(excelName1):
                        wb = load_workbook(excelName1)
                        pass
                    else:
                        wb = Workbook()

                    now_time = time.strftime('%m-%d', time.localtime(time.time()))
                    if wb.sheetnames[0] == str(now_time):
                        ws = wb[str(now_time)]
                        pass
                    else:
                        ws = wb.create_sheet(str(now_time), 0)
                        ws['A1'] = '日期/时间'
                        ws['B1'] = '样品编号'
                        ws['C1'] = '温度'
                        ws['D1'] = '湿度'
                        ws['E1'] = '重量'
                        ws['F1'] = '样品位号'
                    now_time_1 = time.strftime('%Y-%m-%d %H-%M', time.localtime(time.time()))
                    ws.column_dimensions['A'].width = 20
                    ws.column_dimensions['B'].width = 15
                    ws.column_dimensions['C'].width = 15
                    ws.column_dimensions['D'].width = 15
                    ws.column_dimensions['E'].width = 15
                    ws.column_dimensions['F'].width = 15
                    list = [str(now_time_1), str(ban), tem, hum, weight + "g", csnum]
                    # print(list)
                    ws.append(list)
                    wb.save(excelName1)
                    new_name = str(now_time_1) + str(csnum) + ".xlsx"
                    os.rename("text.xlsx", new_name)

                    if os.path.exists(excelName):
                        wb1 = load_workbook(excelName)
                        pass
                    else:
                        wb1 = Workbook()

                    now_time = time.strftime('%m-%d', time.localtime(time.time()))
                    if wb1.sheetnames[0] == str(now_time):
                        ws2 = wb1[str(now_time)]
                        pass
                    else:
                        ws2 = wb1.create_sheet(str(now_time), 0)
                        ws2['A1'] = '日期/时间'
                        ws2['B1'] = '样品编号'
                        ws2['C1'] = '温度'
                        ws2['D1'] = '湿度'
                        ws2['E1'] = '重量'
                        ws2['F1'] = '样品位号'
                    now_time_2 = time.strftime('%Y-%m-%d %H-%M', time.localtime(time.time()))
                    ws2.column_dimensions['A'].width = 20
                    ws2.column_dimensions['B'].width = 15
                    ws2.column_dimensions['C'].width = 15
                    ws2.column_dimensions['D'].width = 15
                    ws2.column_dimensions['E'].width = 15
                    ws2.column_dimensions['F'].width = 15
                    list1 = [str(now_time_2), str(ban), tem, hum, weight + "g", csnum]
                    # print(list)
                    ws2.append(list1)
                    wb1.save(excelName)

                    zmc.axis_3_up()
                    if zmc.get_input(15) == 1:
                        break
                    time.sleep(25)
                    zmc.axis_0_move(8540)
                    if zmc.get_input(15) == 1:
                        break
                    time.sleep(5)
                    zmc.axis_1_up(76000)
                    if zmc.get_input(15) == 1:
                        break
                    time.sleep(6)
                    zmc.axis_0_move(0)
                    if zmc.get_input(15) == 1:
                        break
                    time.sleep(9)
                    zmc.axis_1_up(150000)
                    if zmc.get_input(15) == 1:
                        break
                    time.sleep(20)
                    zmc.axis_0_move(-2642)
                    if zmc.get_input(15) == 1:
                        break
                    time.sleep(4)
                    zmc.axis_1_up(134000)
                    if zmc.get_input(15) == 1:
                        break
                    time.sleep(5)
                    zmc.axis_2_move(8572)
                    if zmc.get_input(15) == 1:
                        break
                    time.sleep(4)
            zmc.axis_0_move(0)
            time.sleep(5)
            zmc.axis_1_back()
            now_time_1 = time.strftime('%Y-%m-%d %H-%M', time.localtime(time.time()))
            self.signal_T1.emit(str(now_time_1))
            try:
                new_name2 = str(now_time_1) + ".xlsx"
                os.rename("result.xlsx", new_name2)
            except:
                time.sleep(1)
            time.sleep(10)
            self.signal_T0.emit('')
            self.signal_T1.emit('')
            # qmut_1.unlock()


    class Thread_2(QThread):  # 恒温称重
        global weight
        signal_I = pyqtSignal(str)
        signal_W = pyqtSignal(str)
        signal_B = pyqtSignal(str)
        signal_T0 = pyqtSignal(str)
        signal_T1 = pyqtSignal(str)

        def __init__(self):
            super().__init__()

        def run(self):
            # qmut_2.lock()
            ip_2048 = "192.168.0.11"
            zmc.connect(ip_2048)  # 连接控制器
            # zmc.read_info()  # 读取控制器信息
            zmc.set_axis_type()
            zmc.set_axis_units()
            time.sleep(7200)
            self.signal_T1.emit('')
            now_time_0 = time.strftime('%Y-%m-%d %H-%M', time.localtime(time.time()))
            self.signal_T0.emit(str(now_time_0))
            # self.signal_T1.emit( )
            zmc.set_op(7, 0)
            zmc.axis_1_up(25000)  # Z轴升高至第一层样品下位置
            time.sleep(8)  # 时间ok
            zmc.axis_0_move(-2642)  # Z轴旋转至第一层样品下位置
            time.sleep(5)  # 时间ok
            for i in range(1, 15):
                print(i)
                if zmc.get_input(10) == 0:
                    time.sleep(2)
                    self.signal_B.emit('')
                    self.signal_W.emit('')
                    self.signal_I.emit('')
                    zmc.axis_2_move(8572)
                    time.sleep(5)
                    continue
                else:
                    csnum = str(chr(8544)) + "-" + str(i)
                    self.signal_I.emit(csnum)
                    self.signal_W.emit('')
                    zmc.axis_1_up(43000)  # Z轴升高，托住样品
                    if zmc.get_input(15) == 1:
                        break
                    time.sleep(6)  # 时间ok
                    zmc.axis_0_move(0)  # Z轴旋转至初始位置
                    if zmc.get_input(15) == 1:
                        break
                    time.sleep(4)
                    zmc.axis_1_up(76000)
                    if zmc.get_input(15) == 1:
                        break
                    time.sleep(9)
                    zmc.set_op(1, 1)  # 离子风开
                    time.sleep(15)  # 离子风吹15S时间
                    zmc.set_op(1, 0)  # 离子风关

                    y = serial.Serial('com7', 115200, timeout=1)  # 扫码
                    myinput = bytes([0X16, 0X54, 0X0D])  # 需要发送"S"
                    y.write(myinput)  # 用write函数向串口发送数据
                    try:
                        myout = y.read(13)  # 提取接收缓冲区中的前7个字节数
                        ban = int(myout)
                        self.signal_B.emit(str(ban))
                        # print(ban)
                        y.close()
                    except:
                        ban = " "
                        # print(ban)
                        y.close()
                    time.sleep(2)
                    zmc.axis_0_move(8540)
                    if zmc.get_input(15) == 1:
                        break
                    time.sleep(10)
                    zmc.axis_1_up(62000)
                    if zmc.get_input(15) == 1:
                        break
                    time.sleep(6)
                    zmc.axis_0_move(12000)
                    if zmc.get_input(15) == 1:
                        break
                    time.sleep(5)
                    zmc.axis_3_down()
                    if zmc.get_input(15) == 1:
                        break
                    time.sleep(40)
                    th = serial.Serial('com6', 2400, timeout=1)  # 温湿度
                    z = serial.Serial('com2', 9600, timeout=1)  # 天平
                    myinput = bytes([0X53, 0X30, 0X31, 0X39, 0X39, 0X41])  # 需要发送的十六进制数据
                    th.write(myinput)  # 用write函数向串口发送数据
                    myinput1 = bytes([0X53, 0X0A, 0X0D])  # 需要发送"S"
                    z.write(myinput1)  # 用write函数向串口发送数据
                    myout = th.read(25)  # 提取接收缓冲区中的前7个字节数
                    datas = ''.join(map(lambda x: ('/x' if len(hex(x)) >= 4 else '/x0') + hex(x)[2:], myout))
                    new_datas = datas[2:].split('/x')  # 由于datas变量中的数据前两

                    tem = (int(new_datas[3], 16) + int(new_datas[4], 16) * 256 + int(new_datas[5], 16) * 65536 + int(
                        new_datas[6],
                        16) * 16777216) / 100
                    hum = (int(new_datas[11], 16) + int(new_datas[12], 16) * 256 + int(new_datas[13], 16) * 65536 + int(
                        new_datas[14],
                        16) * 16777216) / 100
                    myout1 = z.read(25)  # 提取接收缓冲区中的前7个字节数
                    weight = str(myout1[6:14], encoding='utf-8')
                    self.signal_W.emit(weight + 'g')
                    print(weight + "g")
                    th.close()
                    z.close()
                    time.sleep(1)
                    excelName1 = 'text.xlsx'
                    excelName = 'result.xlsx'

                    if os.path.exists(excelName1):
                        wb = load_workbook(excelName1)
                        pass
                    else:
                        wb = Workbook()

                    now_time = time.strftime('%m-%d', time.localtime(time.time()))
                    if wb.sheetnames[0] == str(now_time):
                        ws = wb[str(now_time)]
                        pass
                    else:
                        ws = wb.create_sheet(str(now_time), 0)
                        ws['A1'] = '日期/时间'
                        ws['B1'] = '样品编号'
                        ws['C1'] = '温度'
                        ws['D1'] = '湿度'
                        ws['E1'] = '重量'
                        ws['F1'] = '样品位号'
                    now_time_1 = time.strftime('%Y-%m-%d %H-%M', time.localtime(time.time()))
                    ws.column_dimensions['A'].width = 20
                    ws.column_dimensions['B'].width = 15
                    ws.column_dimensions['C'].width = 15
                    ws.column_dimensions['D'].width = 15
                    ws.column_dimensions['E'].width = 15
                    ws.column_dimensions['F'].width = 15
                    list = [str(now_time_1), str(ban), tem, hum, weight + "g", csnum]
                    # print(list)
                    ws.append(list)
                    wb.save(excelName1)
                    new_name = str(now_time_1) + str(csnum) + ".xlsx"
                    os.rename("text.xlsx", new_name)

                    if os.path.exists(excelName):
                        wb1 = load_workbook(excelName)
                        pass
                    else:
                        wb1 = Workbook()

                    now_time = time.strftime('%m-%d', time.localtime(time.time()))
                    if wb1.sheetnames[0] == str(now_time):
                        ws2 = wb1[str(now_time)]
                        pass
                    else:
                        ws2 = wb1.create_sheet(str(now_time), 0)
                        ws2['A1'] = '日期/时间'
                        ws2['B1'] = '样品编号'
                        ws2['C1'] = '温度'
                        ws2['D1'] = '湿度'
                        ws2['E1'] = '重量'
                        ws2['F1'] = '样品位号'
                    now_time_2 = time.strftime('%Y-%m-%d %H-%M', time.localtime(time.time()))
                    ws2.column_dimensions['A'].width = 20
                    ws2.column_dimensions['B'].width = 15
                    ws2.column_dimensions['C'].width = 15
                    ws2.column_dimensions['D'].width = 15
                    ws2.column_dimensions['E'].width = 15
                    ws2.column_dimensions['F'].width = 15
                    list1 = [str(now_time_2), str(ban), tem, hum, weight + "g", csnum]
                    # print(list)
                    ws2.append(list1)
                    wb1.save(excelName)

                    zmc.axis_3_up()
                    if zmc.get_input(15) == 1:
                        break
                    time.sleep(25)
                    zmc.axis_0_move(8540)
                    if zmc.get_input(15) == 1:
                        break
                    time.sleep(5)
                    zmc.axis_1_up(76000)
                    if zmc.get_input(15) == 1:
                        break
                    time.sleep(6)
                    zmc.axis_0_move(0)
                    if zmc.get_input(15) == 1:
                        break
                    time.sleep(9)
                    zmc.axis_1_up(43000)
                    if zmc.get_input(15) == 1:
                        break
                    time.sleep(9)
                    zmc.axis_0_move(-2642)
                    if zmc.get_input(15) == 1:
                        break
                    time.sleep(5)
                    zmc.axis_1_up(25000)
                    if zmc.get_input(15) == 1:
                        break
                    time.sleep(5)
                    zmc.axis_2_move(8572)
                    if zmc.get_input(15) == 1:
                        break
                    time.sleep(4)
            zmc.axis_0_move(0)
            time.sleep(4)
            zmc.axis_1_up(134000)
            time.sleep(38)
            zmc.axis_0_move(-2642)
            time.sleep(5)
            for i in range(1, 15):
                print(i)
                if zmc.get_input(10) == 0:  # 判断样品位上是否有样品？？？
                    time.sleep(2)
                    self.signal_B.emit('')
                    self.signal_W.emit('')
                    self.signal_I.emit('')
                    zmc.axis_2_move(8572)
                    time.sleep(5)
                    continue
                else:
                    csnum = str(chr(8545)) + "-" + str(i)
                    self.signal_I.emit(csnum)
                    zmc.axis_1_up(150000)  # Z轴升高，托住样品
                    if zmc.get_input(15) == 1:
                        break
                    time.sleep(6)
                    zmc.axis_0_move(0)  # Z轴旋转至初始位置
                    if zmc.get_input(15) == 1:
                        break
                    time.sleep(5)
                    zmc.axis_1_up(76000)
                    if zmc.get_input(15) == 1:
                        break
                    time.sleep(30)
                    zmc.set_op(1, 1)
                    time.sleep(5)
                    zmc.set_op(1, 0)

                    y = serial.Serial('com7', 115200, timeout=1)  # 扫码
                    myinput = bytes([0X16, 0X54, 0X0D])  # 需要发送"S"
                    y.write(myinput)  # 用write函数向串口发送数据
                    try:
                        myout = y.read(13)  # 提取接收缓冲区中的前7个字节数
                        ban = int(myout)
                        self.signal_B.emit(str(ban))
                        print(ban)
                        y.close()
                    except:
                        ban = " "
                        print(ban)
                        y.close()

                    time.sleep(2)
                    zmc.axis_0_move(8540)
                    if zmc.get_input(15) == 1:
                        break
                    time.sleep(9)
                    zmc.axis_1_up(62000)
                    if zmc.get_input(15) == 1:
                        break
                    time.sleep(6)
                    zmc.axis_0_move(12000)
                    if zmc.get_input(15) == 1:
                        break
                    time.sleep(5)
                    zmc.axis_3_down()
                    if zmc.get_input(15) == 1:
                        break
                    time.sleep(60)
                    th = serial.Serial('com6', 2400, timeout=1)
                    z = serial.Serial('com2', 9600, timeout=1)
                    myinput = bytes([0X53, 0X30, 0X31, 0X39, 0X39, 0X41])  # 需要发送的十六进制数据
                    th.write(myinput)  # 用write函数向串口发送数据
                    myinput1 = bytes([0X53, 0X0A, 0X0D])  # 需要发送"S"
                    z.write(myinput1)  # 用write函数向串口发送数据
                    myout = th.read(25)  # 提取接收缓冲区中的前7个字节数
                    datas = ''.join(map(lambda x: ('/x' if len(hex(x)) >= 4 else '/x0') + hex(x)[2:], myout))
                    new_datas = datas[2:].split('/x')  # 由于datas变量中的数据前两

                    tem = (int(new_datas[3], 16) + int(new_datas[4], 16) * 256 + int(new_datas[5], 16) * 65536 + int(
                        new_datas[6],
                        16) * 16777216) / 100
                    hum = (int(new_datas[11], 16) + int(new_datas[12], 16) * 256 + int(new_datas[13], 16) * 65536 + int(
                        new_datas[14],
                        16) * 16777216) / 100
                    myout1 = z.read(25)  # 提取接收缓冲区中的前7个字节数
                    weight = str(myout1[6:14], encoding='utf-8')
                    self.signal_W.emit(weight + 'g')
                    print(weight)
                    th.close()
                    z.close()
                    time.sleep(1)
                    excelName1 = 'text.xlsx'
                    excelName = 'result.xlsx'

                    if os.path.exists(excelName1):
                        wb = load_workbook(excelName1)
                        pass
                    else:
                        wb = Workbook()

                    now_time = time.strftime('%m-%d', time.localtime(time.time()))
                    if wb.sheetnames[0] == str(now_time):
                        ws = wb[str(now_time)]
                        pass
                    else:
                        ws = wb.create_sheet(str(now_time), 0)
                        ws['A1'] = '日期/时间'
                        ws['B1'] = '样品编号'
                        ws['C1'] = '温度'
                        ws['D1'] = '湿度'
                        ws['E1'] = '重量'
                        ws['F1'] = '样品位号'
                    now_time_1 = time.strftime('%Y-%m-%d %H-%M', time.localtime(time.time()))
                    ws.column_dimensions['A'].width = 20
                    ws.column_dimensions['B'].width = 15
                    ws.column_dimensions['C'].width = 15
                    ws.column_dimensions['D'].width = 15
                    ws.column_dimensions['E'].width = 15
                    ws.column_dimensions['F'].width = 15
                    list = [str(now_time_1), str(ban), tem, hum, weight + "g", csnum]
                    # print(list)
                    ws.append(list)
                    wb.save(excelName1)
                    new_name = str(now_time_1) + str(csnum) + ".xlsx"
                    os.rename("text.xlsx", new_name)

                    if os.path.exists(excelName):
                        wb1 = load_workbook(excelName)
                        pass
                    else:
                        wb1 = Workbook()

                    now_time = time.strftime('%m-%d', time.localtime(time.time()))
                    if wb1.sheetnames[0] == str(now_time):
                        ws2 = wb1[str(now_time)]
                        pass
                    else:
                        ws2 = wb1.create_sheet(str(now_time), 0)
                        ws2['A1'] = '日期/时间'
                        ws2['B1'] = '样品编号'
                        ws2['C1'] = '温度'
                        ws2['D1'] = '湿度'
                        ws2['E1'] = '重量'
                        ws2['F1'] = '样品位号'
                    now_time_2 = time.strftime('%Y-%m-%d %H-%M', time.localtime(time.time()))
                    ws2.column_dimensions['A'].width = 20
                    ws2.column_dimensions['B'].width = 15
                    ws2.column_dimensions['C'].width = 15
                    ws2.column_dimensions['D'].width = 15
                    ws2.column_dimensions['E'].width = 15
                    ws2.column_dimensions['F'].width = 15
                    list1 = [str(now_time_2), str(ban), tem, hum, weight + "g", csnum]
                    # print(list)
                    ws2.append(list1)
                    wb1.save(excelName)

                    zmc.axis_3_up()
                    if zmc.get_input(15) == 1:
                        break
                    time.sleep(25)
                    zmc.axis_0_move(8540)
                    if zmc.get_input(15) == 1:
                        break
                    time.sleep(5)
                    zmc.axis_1_up(76000)
                    if zmc.get_input(15) == 1:
                        break
                    time.sleep(6)
                    zmc.axis_0_move(0)
                    if zmc.get_input(15) == 1:
                        break
                    time.sleep(9)
                    zmc.axis_1_up(150000)
                    if zmc.get_input(15) == 1:
                        break
                    time.sleep(30)
                    zmc.axis_0_move(-2642)
                    if zmc.get_input(15) == 1:
                        break
                    time.sleep(4)
                    zmc.axis_1_up(134000)
                    if zmc.get_input(15) == 1:
                        break
                    time.sleep(5)
                    zmc.axis_2_move(8572)
                    if zmc.get_input(15) == 1:
                        break
                    time.sleep(4)
            zmc.axis_0_move(0)
            time.sleep(5)
            zmc.axis_1_back()
            now_time_1 = time.strftime('%Y-%m-%d %H-%M', time.localtime(time.time()))
            self.signal_T1.emit(str(now_time_1))
            try:
                new_name2 = str(now_time_1) + ".xlsx"
                os.rename("result.xlsx", new_name2)
            except:
                time.sleep(1)
            time.sleep(10)
            self.signal_T0.emit('')
            self.signal_T1.emit('')
            # qmut_1.unlock()


    class Thread_3(QThread):  # 停止
        def __init__(self):
            super().__init__()

        def run(self):
            # qmut_3.lock()
            ip_2048 = "192.168.0.11"
            zmc.connect(ip_2048)  # 连接控制器
            # zmc.read_info()  # 读取控制器信息
            zmc.set_axis_type()
            zmc.set_axis_units()
            zmc.set_op(7, 1)
            time.sleep(10000)
            # zmc.set_op(7, 0)
            # qmut_3.unlock()


    class Thread_4(QThread):  # 天平置零
        signal_W = pyqtSignal(str)

        def __init__(self):
            super().__init__()

        def run(self):
            # qmut_4.lock()
            ip_2048 = "192.168.0.11"
            zmc.connect(ip_2048)  # 连接控制器
            # zmc.read_info()  # 读取控制器信息
            zmc.set_axis_type()
            zmc.set_axis_units()
            zmc.set_op(7, 0)
            x = serial.Serial('com2', 9600, timeout=1)
            myinput = bytes([0X5a, 0X0A, 0X0D])  # 需要发送"Z"
            x.write(myinput)  # 用write函数向串口发送数据
            myout = x.read(25)  # 提取接收缓冲区中的前7个字节数
            weight = str(myout[6:14], encoding='utf-8')
            self.signal_W.emit(weight)
            x.close()


    class Thread_5(QThread):  # 2轴开始运行
        def __init__(self):
            super().__init__()

        def run(self):
            ip_2048 = "192.168.0.11"
            zmc.connect(ip_2048)  # 连接控制器
            # zmc.read_info()  # 读取控制器信息
            zmc.set_axis_type()
            zmc.set_axis_units()
            zmc.set_op(7, 0)
            zmc.set_move(2, 1)


    class Thread_6(QThread):  # 2轴停止运行
        def __init__(self):
            super().__init__()

        def run(self):
            ip_2048 = "192.168.0.11"
            zmc.connect(ip_2048)  # 连接控制器
            # zmc.read_info()  # 读取控制器信息
            zmc.set_axis_type()
            zmc.set_axis_units()
            zmc.set_op(7, 0)
            zmc.set_cancel(2, 0)


    class Thread_7(QThread):  # 手动扫码
        signal_B = pyqtSignal(str)

        def __init__(self):
            super().__init__()

        def run(self):
            y = serial.Serial('com7', 115200, timeout=1)
            myinput = bytes([0X16, 0X54, 0X0D])  # 需要发送"S"
            y.write(myinput)  # 用write函数向串口发送数据
            try:
                myout = y.read(13)  # 提取接收缓冲区中的前7个字节数
                ban = int(myout)
                self.signal_B.emit(str(ban))
                time.sleep(5)
                self.signal_B.emit('')
                y.close()
            except:
                ban = " "
                # print(ban)
                y.close()


    class Thread_8(QThread):  # 手动开启离子风
        def __init__(self):
            super().__init__()

        def run(self):
            zmc.set_op(1, 1)  # 离子风开
            time.sleep(5)  # 离子风吹5S时间
            zmc.set_op(1, 0)  # 离子风关


    class Thread_9(QThread):  # 天平罩上升
        def __init__(self):
            super().__init__()

        def run(self):
            ip_2048 = "192.168.0.11"
            zmc.connect(ip_2048)  # 连接控制器
            # zmc.read_info()  # 读取控制器信息
            zmc.set_axis_type()
            zmc.set_axis_units()
            zmc.set_op(7, 0)
            zmc.axis_3_up()


    class Thread_10(QThread):  # 天平罩下降
        def __init__(self):
            super().__init__()

        def run(self):
            ip_2048 = "192.168.0.11"
            zmc.connect(ip_2048)  # 连接控制器
            # zmc.read_info()  # 读取控制器信息
            zmc.set_axis_type()
            zmc.set_axis_units()
            zmc.set_op(7, 0)
            zmc.axis_3_down()


    class Thread_11(QThread):  # 机械臂左旋
        def __init__(self):
            super().__init__()

        def run(self):
            ip_2048 = "192.168.0.11"
            zmc.connect(ip_2048)  # 连接控制器
            # zmc.read_info()  # 读取控制器信息
            zmc.set_axis_type()
            zmc.set_axis_units()
            # zmc.set_op(7, 0)
            zmc.set_move(0, -1)


    class Thread_12(QThread):  # 机械臂右旋
        def __init__(self):
            super().__init__()

        def run(self):
            ip_2048 = "192.168.0.11"
            zmc.connect(ip_2048)  # 连接控制器
            # zmc.read_info()  # 读取控制器信息
            zmc.set_axis_type()
            zmc.set_axis_units()
            # zmc.set_op(7, 0)
            zmc.set_move(0, 1)


    class Thread_13(QThread):  # O轴上升
        def __init__(self):
            super().__init__()

        def run(self):
            ip_2048 = "192.168.0.11"
            zmc.connect(ip_2048)  # 连接控制器
            # zmc.read_info()  # 读取控制器信息
            zmc.set_axis_type()
            zmc.set_axis_units()
            zmc.set_op(7, 0)
            zmc.set_move(1, 1)


    class Thread_14(QThread):  # O轴下降
        def __init__(self):
            super().__init__()

        def run(self):
            ip_2048 = "192.168.0.11"
            zmc.connect(ip_2048)  # 连接控制器
            # zmc.read_info()  # 读取控制器信息
            zmc.set_axis_type()
            zmc.set_axis_units()
            zmc.set_op(7, 0)
            zmc.set_move(1, -1)


    class Thread_15(QThread):  # 2轴回零
        def __init__(self):
            super().__init__()

        def run(self):
            ip_2048 = "192.168.0.11"
            zmc.connect(ip_2048)  # 连接控制器
            # zmc.read_info()  # 读取控制器信息
            zmc.set_axis_type()
            zmc.set_axis_units()
            zmc.set_op(7, 0)
            zmc.axis_2_back()


    class Thread_16(QThread):  # 左右旋转停止
        def __init__(self):
            super().__init__()

        def run(self):
            ip_2048 = "192.168.0.11"
            zmc.connect(ip_2048)  # 连接控制器
            # zmc.read_info()  # 读取控制器信息
            zmc.set_axis_type()
            zmc.set_axis_units()
            zmc.set_op(7, 0)
            zmc.set_cancel(0, 0)


    class Thread_17(QThread):  # 左右旋转复位
        def __init__(self):
            super().__init__()

        def run(self):
            ip_2048 = "192.168.0.11"
            zmc.connect(ip_2048)  # 连接控制器
            # zmc.read_info()  # 读取控制器信息
            zmc.set_axis_type()
            zmc.set_axis_units()
            zmc.set_op(7, 0)
            zmc.set_op(0, 1)


    class Thread_18(QThread):  # 上下运动复位ok
        def __init__(self):
            super().__init__()

        def run(self):
            ip_2048 = "192.168.0.11"
            zmc.connect(ip_2048)  # 连接控制器
            # zmc.read_info()  # 读取控制器信息
            zmc.set_axis_type()
            zmc.set_axis_units()
            zmc.set_op(7, 0)
            zmc.axis_1_back()


    class Thread_19(QThread):  # 上下运动停止ok
        def __init__(self):
            super().__init__()

        def run(self):
            ip_2048 = "192.168.0.11"
            zmc.connect(ip_2048)  # 连接控制器
            # zmc.read_info()  # 读取控制器信息
            zmc.set_axis_type()
            zmc.set_axis_units()
            zmc.set_op(7, 0)
            zmc.set_cancel(1, 0)


    class MyThread(QThread):  # 实时显示温湿度及天平示数
        # global tem, hum, ban
        signal_T = pyqtSignal(str)  # 自定义一个pyqtSignal信号,信号参数是个字符串str类型
        signal_H = pyqtSignal(str)

        # signal_W = pyqtSignal(str)

        def __init__(self):
            super(MyThread, self).__init__()

        def run(self):
            for i in range(10000000):
                th = serial.Serial('com6', 2400, timeout=1)
                myinput = bytes([0X53, 0X30, 0X31, 0X39, 0X39, 0X41])  # 需要发送的十六进制数据
                th.write(myinput)  # 用write函数向串口发送数据
                myout = th.read(25)  # 提取接收缓冲区中的前7个字节数
                datas = ''.join(map(lambda x: ('/x' if len(hex(x)) >= 4 else '/x0') + hex(x)[2:], myout))
                new_datas = datas[2:].split('/x')  # 由于datas变量中的数据前两

                tem = (int(new_datas[3], 16) + int(new_datas[4], 16) * 256 + int(new_datas[5], 16) * 65536 + int(
                    new_datas[6],
                    16) * 16777216) / 100
                hum = (int(new_datas[11], 16) + int(new_datas[12], 16) * 256 + int(new_datas[13], 16) * 65536 + int(
                    new_datas[14],
                    16) * 16777216) / 100
                # print("现在温度：", tem)
                # print("现在湿度：", hum)
                self.signal_T.emit(str(tem))
                self.signal_H.emit(str(hum))

                # y = serial.Serial('com2', 9600, timeout=1)
                # myinput1 = bytes([0X53, 0X0A, 0X0D])  # 需要发送"S"
                # y.write(myinput1)  # 用write函数向串口发送数据
                # myout1 = y.read(25)  # 提取接收缓冲区中的前7个字节数
                # weight = str(myout1[6:14], encoding='utf-8')
                # print("重量：", ban)
                # self.signal_W.emit(str(weight) + 'g')
                if zmc.get_input(9) == 1:  # 天平罩降至底部
                    th.close()
                    # y.close()
                    time.sleep(65)
                else:
                    th.close()
                    # y.close()
                    time.sleep(1)


    if __name__ == "__main__":
        app = QtWidgets.QApplication(sys.argv)
        myshow = Pyqt5_Serial()
        myshow.show()
        sys.exit(app.exec_())
