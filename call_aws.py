# -*- coding: utf-8 -*-
"""
@Time : 2021/12/30 10:41
@Author : Loonghau XU
@FileName: call_UI_gamsyzs.py
@SoftWare: PyCharm
"""
import configparser
import os
import sys
import time

import serial
from PyQt5 import QtCore, QtWidgets
from PyQt5.QtCore import QTimer, pyqtSignal, QThread, Qt
from PyQt5.QtGui import QStatusTipEvent, QIcon, QFont
from PyQt5.QtWidgets import QApplication, QMainWindow, QDesktopWidget, QMessageBox, QLabel, QTableWidgetItem
from openpyxl import load_workbook, Workbook

from UI_aws_date import Ui_Form_date
from UI_aws_help import Ui_Form_help
from UI_aws_logo import Ui_Form_logo
from UI_aws_loobo import Ui_Form_aws
from UI_aws_mac import Ui_Form_mac
from UI_aws_main import Ui_mainWindow
from UI_limit_mac import Ui_macquanxian
from UI_limit_yj import Ui_yjquanxian
from UI_new_object import Ui_Form_ob
from UI_position_set import Ui_Form_position
from UI_system_set import Ui_Form_system
from temp_set import Ui_Form_temset
from zmcwrapper import ZMCWrapper

zmc = ZMCWrapper()  # 实例化类
ip_2048 = "192.168.0.11"

BASE_DIR = os.path.dirname(os.path.realpath(sys.argv[0]))
conf = configparser.RawConfigParser()
conf.read(os.path.join(BASE_DIR, 'setting_aws.ini'))
com_temp = conf.get('COM', 'com_humi')
baud_rate_t = int(conf.get('COM', 'baud_rate_hu'))
try:
    ser_tem = serial.Serial('com' + com_temp, baud_rate_t, timeout=0.8)
    ser_tem.close()
except BaseException:
    pass


class MyMainWindow(QMainWindow, Ui_mainWindow):
    def __init__(self, parent=None):
        super(MyMainWindow, self).__init__(parent)
        self.setupUi(self)
        self.setWindowTitle("全自动恒温恒湿称重系统")  # 窗口名称
        self.setWindowIcon(QIcon('./images/auto_05.ico'))
        self.status = self.statusBar()
        self.status.showMessage("青岛路博建业环保科技有限公司")  # 状态栏信息
        self.kongbai = QLabel()
        self.kongbai.setStyleSheet(
            "min-width:160px;min-height:25px;max-width:20px;max-height:25px;font-size: 15px;")
        self.kongbai.setAlignment(
            QtCore.Qt.AlignRight | QtCore.Qt.AlignVCenter)
        self.status.addPermanentWidget(self.kongbai, stretch=0)
        allow_deviation = conf.get('Parameter', 'allow_deviation')
        self.label_26.setText(allow_deviation)
        self.checkBox.stateChanged.connect(self.state_change)
        self.center()  # 界面居中显示
        self.temp_humi_read()  # 温度读取函数
        self.ipconnect()  # 网络链接确认
        self.slot_def()  # 槽函数

    def slot_def(self):
        self.actionbanben.triggered.connect(self.loobo_logo)
        self.actionbangzhu.triggered.connect(self.help_page)
        self.actiondanbukongzhi.triggered.connect(self.new_control)
        self.actionxitongshezhi.triggered.connect(self.new_control)
        self.action_2.triggered.connect(self.tem_hum_control)  # 温湿度设置按钮
        self.actionguanbi.triggered.connect(self.close)  # 关闭按钮
        self.actionxinjian.triggered.connect(self.new_object)  # 新建功能
        self.pushButton_4.clicked.connect(self.first_station)  # 回零按钮
        self.actionweizhishezhi.triggered.connect(self.new_control)  # 位置设置按钮
        self.pushButton.clicked.connect(self.object_begin)
        self.pushButton_2.clicked.connect(self.object_stop)

    def object_begin(self):
        self.label_35.setText('')
        if zmc.get_DPOS(0) == 0 and zmc.get_DPOS(
                1) == 0 and zmc.get_DPOS(2) == 0 and zmc.get_DPOS(3) == 0:
            if self.label_28.text() == '':
                QMessageBox.warning(
                    self,
                    "项目",
                    "项目未建立",
                    QMessageBox.Yes | QMessageBox.No,
                    QMessageBox.Yes)
                pass
            else:
                self.pushButton.setEnabled(False)
                if self.date_availability():
                    print("20220413")
                    self.timer_delay = QTimer()
                    self.timer_delay.timeout.connect(self.object_begin_time)
                    delay_start = conf.get('Parameter', 'delay_start')
                    self.timer_delay.start(3600000 * float(delay_start))
                else:
                    QMessageBox.warning(
                        self,
                        "程序升级提醒",
                        "程序请升级",
                        QMessageBox.Yes | QMessageBox.No,
                        QMessageBox.Yes)
                    pass
        else:
            QMessageBox.warning(
                self,
                "初始状态",
                "机械手未在初始位",
                QMessageBox.Yes | QMessageBox.No,
                QMessageBox.Yes)
            pass

    def object_begin_time(self):
        self.timer_delay.stop()
        if self.checkBox.isChecked() == False:
            if self.label_28.text() == '低浓度头/滤筒':
                num_sample = 1
                code_recognition = self.label_30.text()
                weight_number = self.label_33.text()
                weight_frequency = self.label_34.text()
                self.thread_weight1 = ThreadWeight(
                    code_recognition, weight_number, weight_frequency, num_sample)
                self.thread_weight1.start()
                self.thread_weight1.signal_layer.connect(self.callback_layer)
                self.thread_weight1.signal_position.connect(
                    self.callback_position)
                self.thread_weight1.signal_wight.connect(self.callback_wight)
                self.thread_weight1.signal_wight2.connect(self.callback_wight2)
                self.thread_weight1.signal_code.connect(self.callback_code)
                self.thread_weight1.signal_T0.connect(self.callback_T0)
                self.thread_weight1.signal_T1.connect(self.callback_T1)
            elif self.label_28.text() == '47mm滤膜':
                num_sample = 2
                code_recognition = self.label_30.text()
                weight_number = self.label_33.text()
                weight_frequency = self.label_34.text()
                self.thread_weight2 = ThreadWeight(
                    code_recognition, weight_number, weight_frequency, num_sample)
                self.thread_weight2.start()
                self.thread_weight2.signal_layer.connect(self.callback_layer)
                self.thread_weight2.signal_position.connect(
                    self.callback_position)
                self.thread_weight2.signal_wight.connect(self.callback_wight)
                self.thread_weight2.signal_wight2.connect(self.callback_wight2)
                self.thread_weight2.signal_code.connect(self.callback_code)
                self.thread_weight2.signal_T0.connect(self.callback_T0)
                self.thread_weight2.signal_T1.connect(self.callback_T1)
            elif self.label_28.text() == '90mm滤膜':
                num_sample = 3
                code_recognition = self.label_30.text()
                weight_number = self.label_33.text()
                weight_frequency = self.label_34.text()
                self.thread_weight3 = ThreadWeight(
                    code_recognition, weight_number, weight_frequency, num_sample)
                self.thread_weight3.start()
                self.thread_weight3.signal_layer.connect(self.callback_layer)
                self.thread_weight3.signal_position.connect(
                    self.callback_position)
                self.thread_weight3.signal_wight.connect(self.callback_wight)
                self.thread_weight3.signal_wight2.connect(self.callback_wight2)
                self.thread_weight3.signal_code.connect(self.callback_code)
                self.thread_weight3.signal_T0.connect(self.callback_T0)
                self.thread_weight3.signal_T1.connect(self.callback_T1)
        else:
            self.judge_tem_hun()

    def judge_tem_hun(self):
        tem_now = self.label_24.text()
        hum_now = self.label_23.text()
        tem_set = int(float(conf.get('Parameter', 'tem_set')))
        hum_set = int(float(conf.get('Parameter', 'hum_set')))
        tem_devia = int(float(conf.get('Parameter', 'tem_devia')))
        hum_devia = int(float(conf.get('Parameter', 'hum_devia')))
        if abs(
            float(tem_now) -
            tem_set) < tem_devia and abs(
                float(hum_now) -
                hum_set) < hum_devia:
            if self.label_28.text() == '低浓度头/滤筒':
                num_sample = 1
                code_recognition = self.label_30.text()
                weight_number = self.label_33.text()
                weight_frequency = self.label_34.text()
                self.thread_weight1 = ThreadWeight(
                    code_recognition, weight_number, weight_frequency, num_sample)
                self.thread_weight1.start()
                self.thread_weight1.signal_layer.connect(self.callback_layer)
                self.thread_weight1.signal_position.connect(
                    self.callback_position)
                self.thread_weight1.signal_wight.connect(self.callback_wight)
                self.thread_weight1.signal_wight2.connect(self.callback_wight2)
                self.thread_weight1.signal_code.connect(self.callback_code)
                self.thread_weight1.signal_T0.connect(self.callback_T0)
                self.thread_weight1.signal_T1.connect(self.callback_T1)
            elif self.label_28.text() == '47mm滤膜':
                num_sample = 2
                code_recognition = self.label_30.text()
                weight_number = self.label_33.text()
                weight_frequency = self.label_34.text()
                self.thread_weight2 = ThreadWeight(
                    code_recognition, weight_number, weight_frequency, num_sample)
                self.thread_weight2.start()
                self.thread_weight2.signal_layer.connect(self.callback_layer)
                self.thread_weight2.signal_position.connect(
                    self.callback_position)
                self.thread_weight2.signal_wight.connect(self.callback_wight)
                self.thread_weight2.signal_wight2.connect(self.callback_wight2)
                self.thread_weight2.signal_code.connect(self.callback_code)
                self.thread_weight2.signal_T0.connect(self.callback_T0)
                self.thread_weight2.signal_T1.connect(self.callback_T1)
            elif self.label_28.text() == '90mm滤膜':
                num_sample = 3
                code_recognition = self.label_30.text()
                weight_number = self.label_33.text()
                weight_frequency = self.label_34.text()
                self.thread_weight3 = ThreadWeight(
                    code_recognition, weight_number, weight_frequency, num_sample)
                self.thread_weight3.start()
                self.thread_weight3.signal_layer.connect(self.callback_layer)
                self.thread_weight3.signal_position.connect(
                    self.callback_position)
                self.thread_weight3.signal_wight.connect(self.callback_wight)
                self.thread_weight3.signal_wight2.connect(self.callback_wight2)
                self.thread_weight3.signal_code.connect(self.callback_code)
                self.thread_weight3.signal_T0.connect(self.callback_T0)
                self.thread_weight3.signal_T1.connect(self.callback_T1)
        else:
            self.thread_weight_wait = ThreadWeightWait()
            self.thread_weight_wait.start()
            self.thread_weight_wait.signal_wait.connect(self.callback_wait)

    def callback_wait(self):
        self.judge_tem_hun()

    def callback_wight(self, weight):  # 对称重数据初始化
        self.label_37.setText(weight)

    def callback_wight2(self, weight, now_time, j, i, k, sample):  # 对称重数据进行存储
        tem_get = self.label_24.text()
        hum_get = self.label_23.text()
        weight_cate = self.label_29.text()
        frequcy = '第' + str(k) + '次称重'
        self.label_37.setText(weight)
        try:
            code_get = self.label_38.text()
        except BaseException:
            code_get = ''
        now_name = time.strftime('%Y-%m-%d', time.localtime(time.time()))
        excel_name = str(now_name) + '.xlsx'
        if os.path.exists(excel_name):
            wb = load_workbook(excel_name)
            pass
        else:
            wb = Workbook()

        now_time00 = time.strftime('%m-%d', time.localtime(time.time()))
        if wb.sheetnames[0] == str(now_time00):
            ws2 = wb[str(now_time00)]
            pass
        else:
            ws2 = wb.create_sheet(str(now_time00), 0)
        ws2['A1'] = '时间'
        ws2['B1'] = '温度'
        ws2['C1'] = '湿度'
        ws2['D1'] = '编码'
        ws2['E1'] = '层号'
        ws2['F1'] = '位号'
        ws2['G1'] = '重量'
        ws2['H1'] = '初重/终重'
        ws2['I1'] = '称重次数'
        ws2['J1'] = '样品种类'
        ws2.column_dimensions['A'].width = 20
        ws2.column_dimensions['B'].width = 8
        ws2.column_dimensions['C'].width = 8
        ws2.column_dimensions['D'].width = 15
        ws2.column_dimensions['E'].width = 8
        ws2.column_dimensions['F'].width = 8
        ws2.column_dimensions['G'].width = 12
        ws2.column_dimensions['H'].width = 12
        ws2.column_dimensions['I'].width = 12
        ws2.column_dimensions['J'].width = 15
        if sample == 1:
            sample_cate = '低浓度头/滤筒'
        elif sample == 2:
            sample_cate = '47mm滤膜'
        elif sample == 3:
            sample_cate = '90mm滤膜'
        list = [
            now_time,
            tem_get,
            hum_get,
            code_get,
            str(j),
            str(i),
            weight,
            weight_cate,
            frequcy,
            sample_cate]
        ws2.append(list)
        wb.save(excel_name)
        row_cut = self.tableWidget.rowCount()
        self.tableWidget.insertRow(row_cut)
        self.tableWidget.horizontalHeader().setFont(QFont('SimSun', 12, QFont.Normal))
        self.tableWidget.verticalHeader().setFont(QFont('Times', 12, QFont.Normal))
        item0 = QTableWidgetItem(now_time)
        item0.setFont(QFont('SimSun', 10, QFont.Normal))
        item0.setTextAlignment(Qt.AlignHCenter | Qt.AlignVCenter)
        self.tableWidget.setItem(row_cut, 0, item0)
        item0 = QTableWidgetItem(tem_get)
        item0.setFont(QFont('SimSun', 10, QFont.Normal))
        item0.setTextAlignment(Qt.AlignHCenter | Qt.AlignVCenter)
        self.tableWidget.setItem(row_cut, 1, item0)
        item0 = QTableWidgetItem(hum_get)
        item0.setFont(QFont('SimSun', 10, QFont.Normal))
        item0.setTextAlignment(Qt.AlignHCenter | Qt.AlignVCenter)
        self.tableWidget.setItem(row_cut, 2, item0)
        item0 = QTableWidgetItem(code_get)
        item0.setFont(QFont('SimSun', 10, QFont.Normal))
        item0.setTextAlignment(Qt.AlignHCenter | Qt.AlignVCenter)
        self.tableWidget.setItem(row_cut, 3, item0)
        item0 = QTableWidgetItem(str(j))
        item0.setFont(QFont('SimSun', 10, QFont.Normal))
        item0.setTextAlignment(Qt.AlignHCenter | Qt.AlignVCenter)
        self.tableWidget.setItem(row_cut, 4, item0)
        item0 = QTableWidgetItem(str(i))
        item0.setFont(QFont('SimSun', 10, QFont.Normal))
        item0.setTextAlignment(Qt.AlignHCenter | Qt.AlignVCenter)
        self.tableWidget.setItem(row_cut, 5, item0)
        item0 = QTableWidgetItem(weight)
        item0.setFont(QFont('SimSun', 10, QFont.Normal))
        item0.setTextAlignment(Qt.AlignHCenter | Qt.AlignVCenter)
        self.tableWidget.setItem(row_cut, 6, item0)
        item0 = QTableWidgetItem(weight_cate)
        item0.setFont(QFont('SimSun', 10, QFont.Normal))
        item0.setTextAlignment(Qt.AlignHCenter | Qt.AlignVCenter)
        self.tableWidget.setItem(row_cut, 7, item0)
        item0 = QTableWidgetItem(frequcy)
        item0.setFont(QFont('SimSun', 10, QFont.Normal))
        item0.setTextAlignment(Qt.AlignHCenter | Qt.AlignVCenter)
        self.tableWidget.setItem(row_cut, 8, item0)
        item0 = QTableWidgetItem(sample_cate)
        item0.setFont(QFont('SimSun', 10, QFont.Normal))
        item0.setTextAlignment(Qt.AlignHCenter | Qt.AlignVCenter)
        self.tableWidget.setItem(row_cut, 9, item0)
        self.tableWidget.setColumnWidth(0, 200)
        self.tableWidget.setColumnWidth(1, 120)
        self.tableWidget.setColumnWidth(2, 150)
        self.tableWidget.setColumnWidth(3, 250)
        self.tableWidget.setColumnWidth(4, 80)
        self.tableWidget.setColumnWidth(5, 80)
        self.tableWidget.setColumnWidth(6, 200)
        self.tableWidget.setColumnWidth(7, 150)
        self.tableWidget.setColumnWidth(8, 200)
        self.tableWidget.setColumnWidth(9, 300)
        self.tableWidget.resizeRowsToContents()
        self.tableWidget.scrollToBottom()

    def callback_layer(self, csnum):
        self.label_40.setText(csnum)
        self.label_37.setText('')
        self.label_38.setText('')
        self.label_39.setText('')

    def callback_position(self, position):
        self.label_39.setText(position)

    def callback_code(self, ban):
        self.label_38.setText(ban)

    def callback_T0(self, now_time_0):
        self.label_36.setText(now_time_0)

    def callback_T1(self, now_time_1):
        self.label_35.setText(now_time_1)
        self.pushButton.setEnabled(True)
        QMessageBox.information(
            self,
            "完成",
            "称重任务已完成",
            QMessageBox.Yes | QMessageBox.No,
            QMessageBox.Yes)

    def object_stop(self):
        zmc.set_op(7, 0)
        sys.exit()

    def state_change(self):
        if self.checkBox.isChecked():
            constant_t_h = True
        else:
            constant_t_h = False
        conf.set('Parameter', 'constant_t_h', constant_t_h)
        with open('setting_aws.ini', 'w') as configfile:
            conf.write(configfile)

    def temp_humi_read(self):
        self.timer_temp_read = QTimer(self)
        self.timer_temp_read.timeout.connect(self.temp_humi_read_Th)
        self.timer_temp_read.start(2000)

    def first_station(self):
        self.first_reset = Thread_Reset()
        self.first_reset.start()
        self.first_reset.signal_ok.connect(self.reset_ok)

    def reset_ok(self):
        QMessageBox.information(
            self,
            "初始化",
            "机械手初始化完成",
            QMessageBox.Yes | QMessageBox.No,
            QMessageBox.Yes)

    def temp_humi_read_Th(self):
        self.mythread_Tem = MyThread_Tem()  # 实例化线程
        self.mythread_Tem.signal_tem_hum.connect(
            self.callback_tem_hum)  # 连接线程类中自定义信号槽到本类的自定义槽函数
        self.mythread_Tem.start()  # 开启线程不是调用run函数而是调用start函数

    def callback_tem_hum(self, tem, hum):
        self.label_24.setText(tem)
        self.label_23.setText(hum)

    def center(self):  # 窗口中间显示
        screen = QDesktopWidget().screenGeometry()
        size = self.geometry()
        newLeft = (screen.width() - size.width()) / 2
        newTop = ((screen.height() - size.height()) / 2) - 40
        self.move(int(newLeft), int(newTop))

    def date_availability(self):
        date = conf.get('ABOUT', 'hex_d')
        date = '0o' + date
        date = str(int(date, 8))
        year = int(date[0:4])
        month = int(date[4:6])
        day = int(date[6:8])
        t = (year, month, day, 0, 0, 0, 0, 0, 0)
        secs = time.mktime(t)
        now_time = time.time()
        if now_time > secs:
            ret = False
        else:
            ret = True
        return ret

    def ipconnect(self):
        try:
            ret = zmc.connect(ip_2048)  # 连接控制器
            # zmc.read_info()  # 读取控制器信息
            if ret == 0:
                zmc.set_axis_type()
                zmc.set_axis_units()
                zmc.set_input()
                zmc.set_op(7, 0)
                zmc.set_op(0, 1)
                zmc.set_op(0, 0)
                zmc.set_DPOS(0, 0)
        except BaseException:
            QMessageBox.information(
                self,
                "主板",
                "主板未连接",
                QMessageBox.Yes | QMessageBox.No,
                QMessageBox.Yes)
        if zmc.is_connected:
            self.kongbai.setText('计算机连接正常')
        else:
            self.kongbai.setText('计算机连接异常')

    def loobo_logo(self):
        self.child = LogoWindow()
        self.child.show()

    def help_page(self):
        self.child5 = HelpWindow()
        self.child5.show()

    def new_object(self):
        self.child4 = WindowsObject()
        self.child4.show()
        self.child4.msg_object.connect(self.object_new)

    def object_new(self, text1, text2, text3, value1, value2):
        self.label_28.setText(text1)
        self.label_29.setText(text2)
        self.label_30.setText(text3)
        self.label_33.setText(str(value1))
        self.label_34.setText(str(value2))

    def new_control(self):
        self.child2 = YingjianQuanXian()
        self.child2.show()

    def tem_hum_control(self):
        self.timer_temp_read.stop()
        self.child3 = Tem_Hum_Set()
        self.child3.show()
        self.child3.signal_open.connect(self.temhum_control_open)
        self.child3.signal_close.connect(self.temhum_control_close)

    def temhum_control_open(self):
        self.timer_temp_read.stop()
        ser_tem.close()

    def temhum_control_close(self):
        ser_tem.close()
        self.timer_temp_read.start(2000)

    def closeEvent(self, event):  # 窗口关闭确认
        reply = QMessageBox.question(
            self,
            '程序退出确认',
            '请确认是否要退出程序？',
            QMessageBox.Yes | QMessageBox.No,
            QMessageBox.No)

        if reply == QMessageBox.Yes:
            event.accept()
        else:
            event.ignore()

    def event(self, QEvent):  # 状态栏长时间显示
        if QEvent.type() == QEvent.StatusTip:
            QEvent = QStatusTipEvent('青岛路博建业环保科技有限公司')
        return super().event(QEvent)


class QThreadDelay(QThread):

    def __init__(self):
        super(QThreadDelay, self).__init__()

    def run(self):
        delay_start = conf.get('Parameter', 'delay_start')
        time.sleep(3600 * float(delay_start))


class ThreadWeightWait(QThread):
    signal_wait = pyqtSignal()

    def __init__(self):
        super(ThreadWeightWait, self).__init__()

    def run(self):
        time.sleep(60)
        self.signal_wait.emit()


class ThreadWeight(QThread):
    signal_layer = pyqtSignal(str)  # 样品层号信号
    signal_position = pyqtSignal(str)  # 样品位号信号
    signal_wight = pyqtSignal(str)  # 重量信号
    signal_wight2 = pyqtSignal(str, str, int, int, int, int)
    signal_code = pyqtSignal(str)  # 编码信号
    signal_T0 = pyqtSignal(str)  # 称重开始时间信号
    signal_T1 = pyqtSignal(str)  # 称重结束时间信号

    def __init__(
            self,
            code_recognition,
            weight_number,
            weight_frequency,
            num_sample):
        super(ThreadWeight, self).__init__()
        self.code_recognition = code_recognition  # 扫码方式  功能完成。☆☆☆☆☆
        self.weight_number = weight_number  # 称重个数
        self.weight_frequency = weight_frequency  # 称重次数，功能完成。☆☆☆☆☆
        self.num_sample = num_sample  # 样品种类

    def run(self):
        ip_2048 = "192.168.0.11"
        zmc.connect(ip_2048)  # 连接控制器
        zmc.set_axis_type()
        zmc.set_axis_units()
        now_time_0 = time.strftime(
            '%Y-%m-%d %H-%M',
            time.localtime(
                time.time()))
        self.signal_T0.emit(str(now_time_0))
        zmc.set_op(7, 0)
        axis1_up = int(conf.get('Position', 'axis1_up'))  # 82000
        axis1_down = int(conf.get('Position', 'axis1_down'))  # 69000
        axis0_simple = int(conf.get('Position', 'axis0_simple'))  # -2557
        axis0_balance = int(conf.get('Position', 'axis0_balance'))  # 8540
        axis0_weigh = int(conf.get('Position', 'axis0_weigh'))  # 12000
        for k in range(int(self.weight_frequency)):  # 控制称重次数
            if self.num_sample == 1:
                starting_layer = int(
                    conf.get(
                        'Filter_ca',
                        'starting_layer'))  # 第一层位置
                layer_spacing = int(
                    conf.get(
                        'Filter_ca',
                        'layer_spacing'))  # 层间距
                layer_number = int(conf.get('Filter_ca', 'layer_number'))  # 层数
            elif self.num_sample == 2:
                starting_layer = int(
                    conf.get(
                        'Filter_me47',
                        'starting_layer'))  # 第一层位置
                layer_spacing = int(
                    conf.get(
                        'Filter_me47',
                        'layer_spacing'))  # 层间距
                layer_number = int(
                    conf.get(
                        'Filter_me47',
                        'layer_number'))  # 层数
            elif self.num_sample == 3:
                starting_layer = int(
                    conf.get(
                        'Filter_me90',
                        'starting_layer'))  # 第一层位置
                layer_spacing = int(
                    conf.get(
                        'Filter_me90',
                        'layer_spacing'))  # 层间距
                layer_number = int(
                    conf.get(
                        'Filter_me90',
                        'layer_number'))  # 层数
            num_weight = 0
            for j in range(int(layer_number)):  # 控制样品层
                self.signal_layer.emit(str(j + 1))
                zmc.axis_1_up(
                    starting_layer +
                    j *
                    layer_spacing)  # Z轴升高至第一层样品下位置
                if j == 0:
                    time.sleep(0.5 + starting_layer / zmc.get_speed(1))
                else:
                    time.sleep(0.5 + layer_spacing / zmc.get_speed(1))
                axis0_simple = int(
                    conf.get(
                        'Position',
                        'axis0_simple'))  # 第一层位置
                zmc.axis_0_move(axis0_simple)  # Z轴旋转至第一层样品下位置
                time.sleep(0.5 + abs(-axis0_simple) / zmc.get_speed(0))
                if self.num_sample == 1:
                    num_rang = 15
                elif self.num_sample == 2:
                    num_rang = 15
                elif self.num_sample == 3:
                    num_rang == 9
                for i in range(1, num_rang):
                    self.signal_wight.emit('')
                    self.signal_code.emit('')
                    if zmc.get_input(15) == 1 or zmc.get_op(7) == 1:
                        break
                    # elif zmc.get_input(10) == 0:  # 判断有无样品
                    #     time.sleep(2)
                    #     self.signal_position.emit('')
                    #     zmc.axis_2_move(8572)
                    #     time.sleep(0.5 + abs(8572) / zmc.get_speed(2))
                    #     continue
                    else:
                        self.signal_position.emit(str(i))
                        now_time = str(
                            time.strftime(
                                '%Y-%m-%d %H-%M',
                                time.localtime(
                                    time.time())))
                        if self.num_sample == 1:
                            up_distance = int(
                                conf.get('Filter_ca', 'up_distance'))
                        elif self.num_sample == 2:
                            up_distance = int(
                                conf.get('Filter_me47', 'up_distance'))
                        elif self.num_sample == 3:
                            up_distance = int(
                                conf.get('Filter_me90', 'up_distance'))
                        zmc.axis_1_up(
                            up_distance +
                            starting_layer +
                            j *
                            layer_spacing)  # Z轴升高，托住样品
                        if zmc.get_input(15) == 1 or zmc.get_op(7) == 1:
                            break
                        time.sleep(
                            0.5 + up_distance / zmc.get_speed(1))  # 时间ok
                        zmc.axis_0_move(0)  # Z轴旋转至初始位置
                        if zmc.get_input(15) == 1 or zmc.get_op(7) == 1:
                            break
                        time.sleep(0.5 + abs(axis0_simple) / zmc.get_speed(0))
                        zmc.axis_1_up(axis1_up)
                        if zmc.get_input(15) == 1 or zmc.get_op(7) == 1:
                            break
                        time.sleep(0.5 +
                                   abs(axis1_up -
                                       (up_distance +
                                        starting_layer +
                                        j *
                                        layer_spacing)) /
                                   zmc.get_speed(1))
                        zmc.set_op(1, 1)  # 离子风开
                        time_ion = int(
                            conf.get(
                                'Parameter',
                                'time_ion'))  # 离子风吹时间
                        time.sleep(time_ion)
                        zmc.set_op(1, 0)  # 离子风关
                        if self.code_recognition == '人工输入':
                            ban = ''
                        else:
                            ban = zmc.code_ban()  # 扫码
                        self.signal_code.emit(str(ban))
                        time.sleep(1)
                        zmc.axis_0_move(axis0_balance)
                        if zmc.get_input(15) == 1 or zmc.get_op(7) == 1:
                            break
                        time.sleep(0.5 + axis0_balance / zmc.get_speed(0))
                        zmc.axis_1_up(axis1_down)
                        if zmc.get_input(15) == 1 or zmc.get_op(7) == 1:
                            break
                        time.sleep(0.5 + (axis1_up - axis1_down) /
                                   zmc.get_speed(1))
                        zmc.axis_0_move(axis0_weigh)
                        if zmc.get_input(15) == 1 or zmc.get_op(7) == 1:
                            break
                        time.sleep(
                            0.5 + (axis0_weigh - axis0_balance) / zmc.get_speed(0))
                        zmc.axis_3_down()
                        if zmc.get_input(15) == 1 or zmc.get_op(7) == 1:
                            break
                        time.sleep(11)
                        time_balance = conf.get('Parameter', 'time_balance')
                        time.sleep(int(time_balance))
                        allow_deviation = conf.get(
                            'Parameter', 'allow_deviation')
                        ban = zmc.weight_ban()
                        if ban == '':
                            self.signal_wight2.emit(
                                ban + 'g', now_time, j + 1, i, k + 1, self.num_sample)
                        else:
                            weight = float(ban) + float(allow_deviation)
                            self.signal_wight2.emit(
                                str(weight) + 'g', now_time, j + 1, i, k + 1, self.num_sample)
                        time.sleep(1)
                        zmc.axis_3_up()
                        if zmc.get_input(15) == 1 or zmc.get_op(7) == 1:
                            break
                        zmc.axis_0_move(axis0_balance)
                        if zmc.get_input(15) == 1 or zmc.get_op(7) == 1:
                            break
                        time.sleep(
                            0.5 + (axis0_weigh - axis0_balance) / zmc.get_speed(0))
                        zmc.axis_1_up(axis1_up)
                        if zmc.get_input(15) == 1 or zmc.get_op(7) == 1:
                            break
                        time.sleep(0.5 + (axis1_up - axis1_down) /
                                   zmc.get_speed(1))
                        zmc.axis_0_move(0)
                        if zmc.get_input(15) == 1 or zmc.get_op(7) == 1:
                            break
                        time.sleep(0.5 + axis0_balance / zmc.get_speed(0))
                        zmc.axis_1_up(
                            up_distance + starting_layer + j * layer_spacing)
                        if zmc.get_input(15) == 1 or zmc.get_op(7) == 1:
                            break
                        time.sleep(0.5 +
                                   abs(axis1_up -
                                       (up_distance +
                                        starting_layer +
                                        j *
                                        layer_spacing)) /
                                   zmc.get_speed(1))
                        zmc.axis_0_move(axis0_simple)
                        if zmc.get_input(15) == 1 or zmc.get_op(7) == 1:
                            break
                        time.sleep(0.5 + abs(axis0_simple) / zmc.get_speed(0))
                        zmc.axis_1_up(starting_layer + j * layer_spacing)
                        if zmc.get_input(15) == 1 or zmc.get_op(7) == 1:
                            break
                        time.sleep(0.5 + up_distance / zmc.get_speed(1))
                        zmc.axis_2_move(120000 / (num_rang - 1))
                        if zmc.get_input(15) == 1 or zmc.get_op(7) == 1:
                            break
                        time.sleep(
                            0.5 + (120000 / (num_rang - 1)) / zmc.get_speed(2))
                        num_weight = num_weight + 1
                        if num_weight > int(self.weight_number) - 1:
                            break
                zmc.axis_0_move(0)
                time.sleep(0.5 + abs(axis0_simple) / zmc.get_speed(0))
                if num_weight > int(self.weight_number) - 1:
                    break
            zmc.set_datumin(2, 7)
            zmc.set_speed(2, 3000)
            zmc.set_single_datum(2, 8)
            zmc.axis_0_move(0)
            time.sleep(0.5 + abs(axis0_simple) / zmc.get_speed(0))
            zmc.axis_1_back()
            time.sleep(0.5 + 34000 / zmc.get_speed(1))
            self.signal_position.emit('')
            self.signal_layer.emit('')
            self.signal_wight.emit('')
            self.signal_code.emit('')
            weight_interval = conf.get('Parameter', 'weight_interval')
            time.sleep(60 * int(weight_interval))
        now_time_1 = time.strftime(
            '%Y-%m-%d %H-%M',
            time.localtime(
                time.time()))
        self.signal_T1.emit(str(now_time_1))


class Thread_weight_beifen(QThread):
    signal_I = pyqtSignal(str)
    signal_W = pyqtSignal(str)
    signal_code = pyqtSignal(str)
    signal_T0 = pyqtSignal(str)  # 开始时间信号
    signal_T1 = pyqtSignal(str)  # 结束时间信号

    def __init__(self):
        super(Thread_weight_beifen, self).__init__()

    def run(self):
        ip_2048 = "192.168.0.11"
        zmc.connect(ip_2048)  # 连接控制器
        zmc.set_axis_type()
        zmc.set_axis_units()
        self.signal_T1.emit('')
        now_time_0 = time.strftime(
            '%Y-%m-%d %H-%M',
            time.localtime(
                time.time()))
        self.signal_T0.emit(str(now_time_0))
        zmc.set_op(7, 0)
        zmc.axis_1_up(34000)  # Z轴升高至第一层样品下位置
        time.sleep(0.5 + 34000 / zmc.get_speed(1))
        zmc.axis_0_move(-2557)  # Z轴旋转至第一层样品下位置
        time.sleep(0.5 + abs(-2557) / zmc.get_speed(0))
        for i in range(1, 15):
            if zmc.get_input(15) == 1 or zmc.get_op(7) == 1:
                break
            elif zmc.get_input(10) == 0:  # 判断有无样品
                time.sleep(2)
                self.signal_code.emit('')
                self.signal_W.emit('')
                self.signal_I.emit('')
                zmc.axis_2_move(8572)
                time.sleep(0.5 + abs(8572) / zmc.get_speed(2))
                continue
            else:
                csnum = str(chr(8544)) + "-" + str(i)
                self.signal_I.emit(csnum)
                self.signal_W.emit('')
                zmc.axis_1_up(53000)  # Z轴升高，托住样品
                if zmc.get_input(15) == 1 or zmc.get_op(7) == 1:
                    break
                time.sleep(0.5 + (53000 - 34000) / zmc.get_speed(1))  # 时间ok
                zmc.axis_0_move(0)  # Z轴旋转至初始位置
                if zmc.get_input(15) == 1 or zmc.get_op(7) == 1:
                    break
                time.sleep(0.5 + 2557 / zmc.get_speed(0))
                zmc.axis_1_up(82000)
                if zmc.get_input(15) == 1 or zmc.get_op(7) == 1:
                    break
                time.sleep(0.5 + (82000 - 53000) / zmc.get_speed(1))
                zmc.set_op(1, 1)  # 离子风开
                time.sleep(5)  # 离子风吹5S时间
                zmc.set_op(1, 0)  # 离子风关

                y = serial.Serial('com30', 115200, timeout=0.5)  # 扫码
                myinput = bytes([0X16, 0X54, 0X0D])  # 需要发送"S"
                y.write(myinput)  # 用write函数向串口发送数据
                try:
                    myout = y.read(13)  # 提取接收缓冲区中的前7个字节数
                    ban = int(myout)
                    self.signal_code.emit(str(ban))
                    y.close()
                except BaseException:
                    ban = " "
                    y.close()

                time.sleep(2)
                zmc.axis_0_move(8540)
                if zmc.get_input(15) == 1 or zmc.get_op(7) == 1:
                    break
                time.sleep(0.5 + 8540 / zmc.get_speed(0))
                zmc.axis_1_up(68000)
                if zmc.get_input(15) == 1 or zmc.get_op(7) == 1:
                    break
                time.sleep(0.5 + (82000 - 68000) / zmc.get_speed(1))
                zmc.axis_0_move(12000)
                if zmc.get_input(15) == 1 or zmc.get_op(7) == 1:
                    break
                time.sleep(0.5 + (12000 - 8540) / zmc.get_speed(0))
                zmc.axis_3_down()
                if zmc.get_input(15) == 1 or zmc.get_op(7) == 1:
                    break
                time.sleep(40)
                # th = serial.Serial('com21', 2400, timeout=1)  # 温湿度
                z = serial.Serial('com29', 9600, timeout=1)  # 天平
                # myinput = bytes([0X53, 0X30, 0X31, 0X39, 0X39, 0X41])  # 需要发送的十六进制数据
                # th.write(myinput)  # 用write函数向串口发送数据
                myinput1 = bytes([0X53, 0X0A, 0X0D])  # 需要发送"S"
                z.write(myinput1)  # 用write函数向串口发送数据
                # myout = th.read(25)  # 提取接收缓冲区中的前7个字节数
                # datas = ''.join(map(lambda x: ('/x' if len(hex(x)) >= 4 else '/x0') + hex(x)[2:], myout))
                # new_datas = datas[2:].split('/x')  # 由于datas变量中的数据前两
                #
                # tem = (int(new_datas[3], 16) + int(new_datas[4], 16) * 256 + int(new_datas[5], 16) * 65536 + int(
                #     new_datas[6],
                #     16) * 16777216) / 100
                # hum = (int(new_datas[11], 16) + int(new_datas[12], 16) * 256 + int(new_datas[13], 16) * 65536 + int(
                #     new_datas[14],
                #     16) * 16777216) / 100
                myout1 = z.read(25)  # 提取接收缓冲区中的前7个字节数
                weight = str(myout1[6:14], encoding='utf-8')
                self.signal_W.emit(weight + 'g')
                print(weight + "g")
                # th.close()
                z.close()
                time.sleep(1)
                # excelName1 = 'text.xlsx'
                # excelName = 'result.xlsx'
                #
                # if os.path.exists(excelName1):
                #     wb = load_workbook(excelName1)
                #     pass
                # else:
                #     wb = Workbook()
                #
                # now_time = time.strftime('%m-%d', time.localtime(time.time()))
                # if wb.sheetnames[0] == str(now_time):
                #     ws = wb[str(now_time)]
                #     pass
                # else:
                #     ws = wb.create_sheet(str(now_time), 0)
                #     ws['A1'] = '日期/时间'
                #     ws['B1'] = '样品编号'
                #     ws['C1'] = '温度'
                #     ws['D1'] = '湿度'
                #     ws['E1'] = '重量'
                #     ws['F1'] = '样品位号'
                # now_time_1 = time.strftime('%Y-%m-%d %H-%M', time.localtime(time.time()))
                # ws.column_dimensions['A'].width = 20
                # ws.column_dimensions['B'].width = 15
                # ws.column_dimensions['C'].width = 15
                # ws.column_dimensions['D'].width = 15
                # ws.column_dimensions['E'].width = 15
                # ws.column_dimensions['F'].width = 15
                # list = [str(now_time_1), str(ban), tem, hum, weight + "g", csnum]
                # # print(list)
                # ws.append(list)
                # wb.save(excelName1)
                # new_name = str(now_time_1) + str(csnum) + ".xlsx"
                # os.rename("text.xlsx", new_name)
                #
                # if os.path.exists(excelName):
                #     wb1 = load_workbook(excelName)
                #     pass
                # else:
                #     wb1 = Workbook()
                #
                # now_time = time.strftime('%m-%d', time.localtime(time.time()))
                # if wb1.sheetnames[0] == str(now_time):
                #     ws2 = wb1[str(now_time)]
                #     pass
                # else:
                #     ws2 = wb1.create_sheet(str(now_time), 0)
                #     ws2['A1'] = '日期/时间'
                #     ws2['B1'] = '样品编号'
                #     ws2['C1'] = '温度'
                #     ws2['D1'] = '湿度'
                #     ws2['E1'] = '重量'
                #     ws2['F1'] = '样品位号'
                # now_time_2 = time.strftime('%Y-%m-%d %H-%M', time.localtime(time.time()))
                # ws2.column_dimensions['A'].width = 20
                # ws2.column_dimensions['B'].width = 15
                # ws2.column_dimensions['C'].width = 15
                # ws2.column_dimensions['D'].width = 15
                # ws2.column_dimensions['E'].width = 15
                # ws2.column_dimensions['F'].width = 15
                # list1 = [str(now_time_2), str(ban), tem, hum, weight + "g", csnum]
                # # print(list)
                # ws2.append(list1)
                # wb1.save(excelName)

                zmc.axis_3_up()
                if zmc.get_input(15) == 1 or zmc.get_op(7) == 1:
                    break
                # time.sleep(25)
                zmc.axis_0_move(8540)
                if zmc.get_input(15) == 1 or zmc.get_op(7) == 1:
                    break
                time.sleep(0.5 + (12000 - 8540) / zmc.get_speed(0))
                zmc.axis_1_up(82000)
                if zmc.get_input(15) == 1 or zmc.get_op(7) == 1:
                    break
                time.sleep(0.5 + (82000 - 68000) / zmc.get_speed(1))
                zmc.axis_0_move(0)
                if zmc.get_input(15) == 1 or zmc.get_op(7) == 1:
                    break
                time.sleep(0.5 + 8540 / zmc.get_speed(0))
                zmc.axis_1_up(53000)
                if zmc.get_input(15) == 1 or zmc.get_op(7) == 1:
                    break
                time.sleep(0.5 + (82000 - 53000) / zmc.get_speed(1))
                zmc.axis_0_move(-2557)
                if zmc.get_input(15) == 1 or zmc.get_op(7) == 1:
                    break
                time.sleep(0.5 + 2557 / zmc.get_speed(0))
                zmc.axis_1_up(34000)
                if zmc.get_input(15) == 1 or zmc.get_op(7) == 1:
                    break
                time.sleep(0.5 + (53000 - 34000) / zmc.get_speed(1))
                zmc.axis_2_move(8572)
                if zmc.get_input(15) == 1 or zmc.get_op(7) == 1:
                    break
                time.sleep(0.5 + 8572 / zmc.get_speed(2))
        zmc.axis_0_move(0)
        time.sleep(4)
        zmc.axis_1_up(134000)
        time.sleep(38)
        zmc.axis_0_move(-2642)
        time.sleep(5)
        for i in range(1, 15):
            print(i)
            if zmc.get_input(15) == 1 or zmc.get_op(7) == 1:
                break
            elif zmc.get_input(10) == 0:
                time.sleep(2)
                self.signal_code.emit('')
                self.signal_W.emit('')
                self.signal_I.emit('')
                zmc.axis_2_move(8572)
                time.sleep(5)
                continue
            else:
                csnum = str(chr(8545)) + "-" + str(i)
                self.signal_I.emit(csnum)
                zmc.axis_1_up(150000)  # Z轴升高，托住样品
                if zmc.get_input(15) == 1 or zmc.get_op(7) == 1:
                    break
                time.sleep(6)
                zmc.axis_0_move(0)  # Z轴旋转至初始位置
                if zmc.get_input(15) == 1 or zmc.get_op(7) == 1:
                    break
                time.sleep(5)
                zmc.axis_1_up(50000)
                if zmc.get_input(15) == 1 or zmc.get_op(7) == 1:
                    break
                time.sleep(30)
                zmc.set_op(1, 1)
                time.sleep(5)
                zmc.set_op(1, 0)

                y = serial.Serial('com30', 115200, timeout=1)  # 扫码
                myinput = bytes([0X16, 0X54, 0X0D])  # 需要发送"S"
                y.write(myinput)  # 用write函数向串口发送数据
                try:
                    myout = y.read(13)  # 提取接收缓冲区中的前7个字节数
                    ban = int(myout)
                    self.signal_code.emit(str(ban))
                    print(ban)
                    y.close()
                except BaseException:
                    ban = " "
                    print(ban)
                    y.close()

                time.sleep(2)
                zmc.axis_0_move(7627)
                if zmc.get_input(15) == 1 or zmc.get_op(7) == 1:
                    break
                time.sleep(9)
                zmc.axis_1_up(35000)
                if zmc.get_input(15) == 1 or zmc.get_op(7) == 1:
                    break
                time.sleep(6)
                zmc.axis_0_move(11000)
                if zmc.get_input(15) == 1 or zmc.get_op(7) == 1:
                    break
                time.sleep(5)
                zmc.axis_3_down()
                if zmc.get_input(15) == 1 or zmc.get_op(7) == 1:
                    break
                time.sleep(60)
                th = serial.Serial('com21', 2400, timeout=1)
                z = serial.Serial('com29', 9600, timeout=1)
                # 需要发送的十六进制数据
                myinput = bytes([0X53, 0X30, 0X31, 0X39, 0X39, 0X41])
                th.write(myinput)  # 用write函数向串口发送数据
                myinput1 = bytes([0X53, 0X0A, 0X0D])  # 需要发送"S"
                z.write(myinput1)  # 用write函数向串口发送数据
                myout = th.read(25)  # 提取接收缓冲区中的前7个字节数
                datas = ''.join(
                    map(lambda x: ('/x' if len(hex(x)) >= 4 else '/x0') + hex(x)[2:], myout))
                new_datas = datas[2:].split('/x')  # 由于datas变量中的数据前两

                tem = (int(new_datas[3],
                           16) + int(new_datas[4],
                                     16) * 256 + int(new_datas[5],
                                                     16) * 65536 + int(new_datas[6],
                                                                       16) * 16777216) / 100
                hum = (int(new_datas[11],
                           16) + int(new_datas[12],
                                     16) * 256 + int(new_datas[13],
                                                     16) * 65536 + int(new_datas[14],
                                                                       16) * 16777216) / 100
                myout1 = z.read(25)  # 提取接收缓冲区中的前7个字节数
                weight = str(myout1[6:14], encoding='utf-8')
                self.signal_W.emit(weight + 'g')
                print(weight)
                th.close()
                z.close()
                time.sleep(1)
                excelName1 = 'text.xlsx'
                excelName = 'result.xlsx'

                if os.path.exists(excelName1):
                    wb = load_workbook(excelName1)
                    pass
                else:
                    wb = Workbook()

                now_time = time.strftime('%m-%d', time.localtime(time.time()))
                if wb.sheetnames[0] == str(now_time):
                    ws = wb[str(now_time)]
                    pass
                else:
                    ws = wb.create_sheet(str(now_time), 0)
                    ws['A1'] = '日期/时间'
                    ws['B1'] = '样品编号'
                    ws['C1'] = '温度'
                    ws['D1'] = '湿度'
                    ws['E1'] = '重量'
                    ws['F1'] = '样品位号'
                now_time_1 = time.strftime(
                    '%Y-%m-%d %H-%M',
                    time.localtime(
                        time.time()))
                ws.column_dimensions['A'].width = 20
                ws.column_dimensions['B'].width = 15
                ws.column_dimensions['C'].width = 15
                ws.column_dimensions['D'].width = 15
                ws.column_dimensions['E'].width = 15
                ws.column_dimensions['F'].width = 15
                list = [
                    str(now_time_1),
                    str(ban),
                    tem,
                    hum,
                    weight + "g",
                    csnum]
                # print(list)
                ws.append(list)
                wb.save(excelName1)
                new_name = str(now_time_1) + str(csnum) + ".xlsx"
                os.rename("text.xlsx", new_name)

                if os.path.exists(excelName):
                    wb1 = load_workbook(excelName)
                    pass
                else:
                    wb1 = Workbook()

                now_time = time.strftime('%m-%d', time.localtime(time.time()))
                if wb1.sheetnames[0] == str(now_time):
                    ws2 = wb1[str(now_time)]
                    pass
                else:
                    ws2 = wb1.create_sheet(str(now_time), 0)
                    ws2['A1'] = '日期/时间'
                    ws2['B1'] = '样品编号'
                    ws2['C1'] = '温度'
                    ws2['D1'] = '湿度'
                    ws2['E1'] = '重量'
                    ws2['F1'] = '样品位号'
                now_time_2 = time.strftime(
                    '%Y-%m-%d %H-%M',
                    time.localtime(
                        time.time()))
                ws2.column_dimensions['A'].width = 20
                ws2.column_dimensions['B'].width = 15
                ws2.column_dimensions['C'].width = 15
                ws2.column_dimensions['D'].width = 15
                ws2.column_dimensions['E'].width = 15
                ws2.column_dimensions['F'].width = 15
                list1 = [
                    str(now_time_2),
                    str(ban),
                    tem,
                    hum,
                    weight + "g",
                    csnum]
                # print(list)
                ws2.append(list1)
                wb1.save(excelName)

                zmc.axis_3_up()
                if zmc.get_input(15) == 1 or zmc.get_op(7) == 1:
                    break
                time.sleep(25)
                zmc.axis_0_move(7627)
                if zmc.get_input(15) == 1 or zmc.get_op(7) == 1:
                    break
                time.sleep(5)
                zmc.axis_1_up(50000)
                if zmc.get_input(15) == 1 or zmc.get_op(7) == 1:
                    break
                time.sleep(6)
                zmc.axis_0_move(0)
                if zmc.get_input(15) == 1 or zmc.get_op(7) == 1:
                    break
                time.sleep(9)
                zmc.axis_1_up(150000)
                if zmc.get_input(15) == 1 or zmc.get_op(7) == 1:
                    break
                time.sleep(30)
                zmc.axis_0_move(-2642)
                if zmc.get_input(15) == 1 or zmc.get_op(7) == 1:
                    break
                time.sleep(4)
                zmc.axis_1_up(134000)
                if zmc.get_input(15) == 1 or zmc.get_op(7) == 1:
                    break
                time.sleep(5)
                zmc.axis_2_move(8572)
                if zmc.get_input(15) == 1 or zmc.get_op(7) == 1:
                    break
                time.sleep(4)
        zmc.axis_0_move(0)
        time.sleep(5)
        zmc.axis_1_back()
        now_time_1 = time.strftime(
            '%Y-%m-%d %H-%M',
            time.localtime(
                time.time()))
        self.signal_T1.emit(str(now_time_1))
        try:
            new_name2 = str(now_time_1) + ".xlsx"
            os.rename("result.xlsx", new_name2)
        except BaseException:
            time.sleep(1)
        time.sleep(10)
        self.signal_T0.emit('')
        self.signal_T1.emit('')


class WindowsObject(QtWidgets.QWidget, Ui_Form_ob):
    msg_object = pyqtSignal(str, str, str, int, int)

    def __init__(self, parent=None):
        super(WindowsObject, self).__init__(parent)
        self.setupUi(self)
        self.comboBox.currentIndexChanged.connect(self.change_max)
        self.pushButton.clicked.connect(self.cancel)
        self.pushButton_2.clicked.connect(self.confirm)
        self.pushButton_3.clicked.connect(self.reset)

    def change_max(self):
        if self.comboBox.currentText() == '低浓度头/滤筒':
            layer_number = int(conf.get('Filter_ca', 'layer_number'))
            self.spinBox_3.setMaximum(14 * layer_number)
        elif self.comboBox.currentText() == '47mm滤膜':
            layer_number = int(conf.get('Filter_me47', 'layer_number'))
            self.spinBox_3.setMaximum(14 * layer_number)
        elif self.comboBox.currentText() == '90mm滤膜':
            layer_number = int(conf.get('Filter_me90', 'layer_number'))
            self.spinBox_3.setMaximum(8 * layer_number)

    def confirm(self):
        text1 = self.comboBox.currentText()
        text2 = self.comboBox_2.currentText()
        text3 = self.comboBox_3.currentText()
        value1 = self.spinBox_3.value()
        value2 = self.spinBox_4.value()
        self.msg_object.emit(text1, text2, text3, value1, value2)
        self.close()

    def cancel(self):
        self.close()

    def reset(self):
        self.msg_object.emit('', '', '', 0, 0)


class MyThread_Tem(QThread):
    global tem, hum
    signal_tem_hum = pyqtSignal(str, str)  # 自定义一个pyqtSignal信号,信号参数是个字符串str类型

    def __init__(self):
        super(MyThread_Tem, self).__init__()

    def run(self):
        try:
            if ser_tem.isOpen():
                ser_tem.close()
            ser_tem.open()
            tem_set = 'S0199A'
            myinput = bytes(tem_set.encode('utf-8'))
            ser_tem.write(myinput)  # 用write函数向串口发送数据
            ser_tem.flushInput()
            myout = ser_tem.read(25)  # 提取接收缓冲区中的前7个字节数
            ser_tem.flushOutput()
            datas = ''.join(
                map(lambda x: ('/x' if len(hex(x)) >= 4 else '/x0') + hex(x)[2:], myout))
            new_datas = datas[2:].split('/x')  # 由于datas变量中的数据前两
            tem = (int(new_datas[3],
                       16) + int(new_datas[4],
                                 16) * 256 + int(new_datas[5],
                                                 16) * 65536 + int(new_datas[6],
                                                                   16) * 16777216) / 100
            hum = (int(new_datas[11],
                       16) + int(new_datas[12],
                                 16) * 256 + int(new_datas[13],
                                                 16) * 65536 + int(new_datas[14],
                                                                   16) * 16777216) / 100
            tem = format(tem, '.1f')
            hum = format(hum, '.1f')
            self.signal_tem_hum.emit(str(tem), str(hum))
            ser_tem.close()
        except BaseException:
            self.signal_tem_hum.emit('', '')
            # ser_tem.close()


class HelpWindow(QtWidgets.QWidget, Ui_Form_help):
    def __init__(self, parent=None):
        super(HelpWindow, self).__init__(parent)
        self.setupUi(self)


class LogoWindow(QtWidgets.QWidget, Ui_Form_logo):
    def __init__(self, parent=None):
        super(LogoWindow, self).__init__(parent)
        self.setupUi(self)
        code_num = conf.get('ABOUT', 'code_number')
        self.label_8.setText(code_num)

        self.pushButton.clicked.connect(self.mac_setting)

    def mac_setting(self):
        self.close()
        self.mac_setting = MacSetWindows()
        self.mac_setting.show()


class MacSetWindows(QtWidgets.QWidget, Ui_macquanxian):
    def __init__(self, parent=None):
        super(MacSetWindows, self).__init__(parent)
        self.setupUi(self)
        self.lineEdit_2.setFocus()
        self.pushButton.clicked.connect(self.confirm)
        self.pushButton.setShortcut("enter")  # 绑定快捷键
        self.pushButton_2.clicked.connect(self.close)
        self.pushButton_2.setShortcut("esc")

    def confirm(self):
        temp_set_conf = conf.get('Login', 'mac_set')
        date_set_conf = conf.get('Login', 'set_hex')
        if self.lineEdit_2.text() == temp_set_conf:
            self.close()
            self.setting = MacWindows()
            self.setting.show()
        elif self.lineEdit_2.text() == date_set_conf:
            self.close()
            self.setting_d = DateWindows()
            self.setting_d.show()

        else:
            QMessageBox.warning(self,
                                "警告",
                                "用户名或密码错误！",
                                QMessageBox.Yes)
            self.lineEdit_2.end(True)  # 定位光标到最右边


class MacWindows(QtWidgets.QWidget, Ui_Form_mac):
    def __init__(self, parent=None):
        super(MacWindows, self).__init__(parent)
        self.setupUi(self)
        code_num = conf.get('ABOUT', 'code_number')

        self.lineEdit_7.setText(code_num)

        self.pushButton.clicked.connect(self.mac_add_set)
        self.pushButton.setShortcut('enter')
        self.pushButton_2.clicked.connect(self.close)
        self.pushButton_2.setShortcut('esc')

    def mac_add_set(self):
        try:
            mac = self.lineEdit.text()
        except BaseException:
            mac = ''
        try:
            com_humiture = self.lineEdit_2.text()
        except BaseException:
            com_humiture = ''
        try:
            com_balance = self.lineEdit_3.text()
        except BaseException:
            com_balance = ''
        try:
            com_code = self.lineEdit_4.text()
        except BaseException:
            com_code = ''
        try:
            com_motor = self.lineEdit_6.text()
        except BaseException:
            com_motor = ''
        try:
            code_num = self.lineEdit_7.text()
        except BaseException:
            code_num = ''
        try:
            baud_hum = self.lineEdit_8.text()
        except BaseException:
            baud_hum = ''
        try:
            baud_balan = self.lineEdit_9.text()
        except BaseException:
            baud_balan = ''
        try:
            baud_code = self.lineEdit_10.text()
        except BaseException:
            baud_code = ''
        try:
            baud_motor = self.lineEdit_12.text()
        except BaseException:
            baud_motor = ''

        if mac != '':
            conf.set('ABOUT', 'mac_address', mac)
        if com_humiture != '':
            conf.set('COM', 'com_humi', com_humiture)
        if com_balance != '':
            conf.set('COM', 'com_balance', com_balance)
        if com_code != '':
            conf.set('COM', 'com_code', com_code)
        if com_motor != '':
            conf.set('COM', 'com_motor', com_motor)
        if baud_hum != '':
            conf.set('COM', 'baud_rate_hu', baud_hum)
        if baud_balan != '':
            conf.set('COM', 'baud_rate_ba', baud_balan)
        if baud_code != '':
            conf.set('COM', 'baud_rate_co', baud_code)
        if baud_motor != '':
            conf.set('COM', 'baud_rate_mo', baud_motor)
        if code_num != '':
            conf.set('ABOUT', 'code_number', code_num)

        with open('setting_aws.ini', 'w') as configfile:
            conf.write(configfile)
        self.close()


class DateWindows(QtWidgets.QWidget, Ui_Form_date):
    def __init__(self, parent=None):
        super(DateWindows, self).__init__(parent)
        self.setupUi(self)
        self.pushButton.clicked.connect(self.date_add_set)
        self.pushButton.setShortcut('enter')
        self.pushButton_2.clicked.connect(self.close)
        self.pushButton_2.setShortcut('esc')

    def date_add_set(self):
        try:
            date_avai = self.dateEdit.date().toString("yyyyMMdd")
        except BaseException:
            date_avai = ''

        if date_avai != '':
            date_avai = int(date_avai)
            date_avai = oct(date_avai)
            date_avai = date_avai[2:]
            conf.set('ABOUT', 'hex_d', date_avai)

        with open('setting_aws.ini', 'w') as configfile:
            conf.write(configfile)
        self.close()


class Thread_Reset(QThread):
    signal_ok = pyqtSignal()

    def __init__(self):
        super(Thread_Reset, self).__init__()

    def run(self):
        zmc.set_axis_type()
        zmc.set_axis_units()
        zmc.set_op(7, 0)
        zmc.set_speed(2, 3000)
        zmc.set_single_datum(2, 8)
        if zmc.get_input(8) != 1:
            zmc.set_speed(3, 5000)
            zmc.set_single_datum(3, 8)
            time.sleep(3)
        else:
            pass
        if zmc.get_DPOS(0) > 1200:
            zmc.set_singelmoveabs(1, 82000)
            time.sleep(3)
        else:
            pass
        zmc.set_op(0, 1)
        zmc.set_op(0, 0)
        zmc.axis_1_back()
        for i in range(60):
            if zmc.get_DPOS(0) == 0 and zmc.get_DPOS(
                    1) == 0 and zmc.get_DPOS(2) == 0 and zmc.get_DPOS(3) == 0:
                self.signal_ok.emit()
                break
            else:
                time.sleep(2)
                continue


class Tem_Hum_Set(QtWidgets.QWidget, Ui_Form_temset):
    signal_open = pyqtSignal()
    signal_close = pyqtSignal()

    def __init__(self, parent=None):
        super(Tem_Hum_Set, self).__init__(parent)
        self.setupUi(self)
        self.signal_open.emit()
        self.temhum_set_read()
        self.time_temp = QTimer()
        self.time_temp.timeout.connect(self.temp_humi)
        self.time_temp.start(2000)
        self.pushButton_3.clicked.connect(self.tem_set)
        self.pushButton_4.clicked.connect(self.hum_set)

    def temhum_set_read(self):
        if ser_tem.isOpen():
            ser_tem.close()
        self.mythread_read = MyThread_Tem_Hum_Set()
        self.mythread_read.start()
        self.mythread_read.signal_read.connect(self.temhum_read)

    def temp_humi(self):
        self.mythread = MyThread_Tem()  # 实例化线程
        self.mythread.signal_tem_hum.connect(
            self.callback_Blance)  # 连接线程类中自定义信号槽到本类的自定义槽函数
        self.mythread.start()  # 开启线程不是调用run函数而是调用start函数

    def tem_set(self):
        self.time_temp.stop()
        ser_tem.close()
        temset = self.doubleSpinBox_2.value()
        temset = format(temset, '.2f')
        self.mythread_tem_set = MyThread_TemSet(temset)
        self.mythread_tem_set.start()
        self.mythread_tem_set.msg_temset.connect(self.time_qtimer)

    def hum_set(self):
        self.time_temp.stop()
        ser_tem.close()
        humset = self.doubleSpinBox.value()
        humset = format(humset, '.1f')
        self.mythread_hum_set = MyThread_HumSet(humset)
        self.mythread_hum_set.start()
        self.mythread_hum_set.msg_humset.connect(self.time_qtimer)

    def start(self, start_stop):
        if start_stop:
            self.time_temp.stop()
            ser_tem.close()
            self.mythread_start = MyThread_Start()  # 实例化线程
            self.mythread_start.start()  # 开启线程不是调用run函数而是调用start函数
            self.mythread_start.msg_start.connect(self.time_qtimer)
            self.pushButton.setText("停止")
        else:
            self.time_temp.stop()
            ser_tem.close()
            self.mythread_stop = MyThread_Stop()  # 实例化线程
            self.mythread_stop.start()  # 开启线程不是调用run函数而是调用start函数
            self.mythread_stop.msg_stop.connect(self.time_qtimer)
            self.pushButton.setText("启动")

    def stop(self, start_stop):
        if start_stop:
            self.time_temp.stop()
            ser_tem.close()
            self.mythread_stop = MyThread_Stop()  # 实例化线程
            self.mythread_stop.start()  # 开启线程不是调用run函数而是调用start函数
            self.mythread_stop.msg_stop.connect(self.time_qtimer)
            self.pushButton.setText("启动")
        else:
            self.time_temp.stop()
            ser_tem.close()
            self.mythread_start = MyThread_Start()  # 实例化线程
            self.mythread_start.start()  # 开启线程不是调用run函数而是调用start函数
            self.mythread_start.msg_start.connect(self.time_qtimer)
            self.pushButton.setText("停止")

    def time_qtimer(self):
        self.time_temp.start(2000)

    def callback_Blance(self, tem, hum):
        self.label.setText(tem)
        self.label_3.setText(hum)

    def temhum_read(self, tem_set, hum_set, station):
        if tem_set != '':
            self.doubleSpinBox_2.setValue(float(tem_set))
            self.doubleSpinBox.setValue(float(hum_set))
            self.time_temp.start(1000)
            if int(station) == 3:
                self.pushButton.setText("停止")
                self.pushButton.clicked[bool].connect(self.stop)
            elif int(station) == 2:
                self.pushButton.setText("启动")
                self.pushButton.clicked[bool].connect(self.start)
        else:
            pass

    def closeEvent(self, event):
        reply = QMessageBox.question(
            self,
            '窗口',
            '是否要关闭窗口',
            QMessageBox.Yes | QMessageBox.No,
            QMessageBox.No)

        if reply == QMessageBox.Yes:
            event.accept()
            self.time_temp.stop()
            self.signal_close.emit()
        else:
            event.ignore()


class MyThread_Start(QThread):
    msg_start = pyqtSignal()

    def __init__(self):
        super(MyThread_Start, self).__init__()

    def run(self):
        if ser_tem.isOpen():
            ser_tem.close()
        ser_tem.open()
        tem_set = 'E01#99#1#A'
        myinput = bytes(tem_set.encode('utf-8'))
        ser_tem.write(myinput)  # 用write函数向串口发送数据
        ser_tem.flushInput()
        ser_tem.read(1)
        ser_tem.flushOutput()
        ser_tem.close()
        self.msg_start.emit()


class MyThread_Stop(QThread):
    msg_stop = pyqtSignal()

    def __init__(self):
        super(MyThread_Stop, self).__init__()

    def run(self):
        if ser_tem.isOpen():
            ser_tem.close()
        ser_tem.open()
        tem_set = 'E01#99#2#A'
        myinput = bytes(tem_set.encode('utf-8'))
        ser_tem.write(myinput)  # 用write函数向串口发送数据
        ser_tem.flushInput()
        ser_tem.read(1)
        ser_tem.flushOutput()
        ser_tem.close()
        self.msg_stop.emit()


class MyThread_TemSet(QThread):
    msg_temset = pyqtSignal()

    def __init__(self, temset):
        super(MyThread_TemSet, self).__init__()
        self.temset = temset

    def run(self):
        tem_set = str(self.temset)
        tem_set = 'E01#62#01#02#' + tem_set + '#A'
        myinput = bytes(tem_set.encode('utf-8'))
        if ser_tem.isOpen():
            ser_tem.close()
        ser_tem.open()
        ser_tem.write(myinput)  # 用write函数向串口发送数据
        ser_tem.flushInput()
        ser_tem.read(1)
        ser_tem.flushOutput()
        ser_tem.close()
        self.msg_temset.emit()


class MyThread_HumSet(QThread):
    msg_humset = pyqtSignal()

    def __init__(self, humset):
        super(MyThread_HumSet, self).__init__()
        self.humset = humset

    def run(self):
        humset = str(self.humset)
        hum_set = 'E01#62#02#02#' + humset + '#A'
        myinput = bytes(hum_set.encode('utf-8'))
        if ser_tem.isOpen():
            ser_tem.close()
        ser_tem.open()
        ser_tem.write(myinput)  # 用write函数向串口发送数据
        ser_tem.flushInput()
        ser_tem.read(1)
        ser_tem.flushOutput()
        ser_tem.close()
        self.msg_humset.emit()


class MyThread_Tem_Hum_Set(QThread):
    signal_read = pyqtSignal(str, str, str)  # 自定义一个pyqtSignal信号,信号参数是个字符串str类型

    def __init__(self):
        super(MyThread_Tem_Hum_Set, self).__init__()

    def run(self):
        try:
            if ser_tem.isOpen():
                ser_tem.close()
            ser_tem.open()
            tem_set = 'S0199A'
            myinput = bytes(tem_set.encode('utf-8'))
            ser_tem.write(myinput)  # 用write函数向串口发送数据
            ser_tem.flushInput()
            myout = ser_tem.read(20)  # 提取接收缓冲区中的前7个字节数
            ser_tem.flushOutput()
            datas = ''.join(
                map(lambda x: ('/x' if len(hex(x)) >= 4 else '/x0') + hex(x)[2:], myout))
            new_datas = datas[2:].split('/x')  # 由于datas变量中的数据前两
            tem_set = (int(new_datas[7],
                           16) + int(new_datas[8],
                                     16) * 256 + int(new_datas[9],
                                                     16) * 65536 + int(new_datas[10],
                                                                       16) * 16777216) / 100
            hum_set = (int(new_datas[15],
                           16) + int(new_datas[16],
                                     16) * 256 + int(new_datas[17],
                                                     16) * 65536 + int(new_datas[18],
                                                                       16) * 16777216) / 100
            station = int(new_datas[2], 16)
            tem_set = format(tem_set, '.2f')
            hum_set = format(hum_set, '.1f')
            self.signal_read.emit(str(tem_set), str(hum_set), str(station))
            ser_tem.close()
        except BaseException:
            self.signal_read.emit('', '', '')
            ser_tem.close()


class YingjianQuanXian(QtWidgets.QWidget, Ui_yjquanxian):

    def __init__(self, parent=None):
        super(YingjianQuanXian, self).__init__(parent)
        self.setupUi(self)

        self.lineEdit_2.setFocus()
        self.pushButton.clicked.connect(self.confirm)
        self.pushButton.setShortcut("enter")  # 绑定快捷键
        self.pushButton_2.clicked.connect(self.close)
        self.pushButton_2.setShortcut("esc")

    def confirm(self):
        hardware_set_conf = conf.get('Login', 'hardware_set')
        system_set_config = conf.get('Login', 'tem_set')
        position_set_config = conf.get('Login', 'position_set')

        if self.lineEdit_2.text() == hardware_set_conf:
            self.close()
            self.yingjian = YingjianWindow()
            self.yingjian.show()
        elif self.lineEdit_2.text() == system_set_config:
            self.close()
            self.system_set = SetWindows()
            self.system_set.show()
        elif self.lineEdit_2.text() == position_set_config:
            self.close()
            self.position_set = PositionWindows()
            self.position_set.show()
        else:
            QMessageBox.warning(self,
                                "警告",
                                "用户名或密码错误！",
                                QMessageBox.Yes)
            self.lineEdit_2.end(True)  # 定位光标到最右边


class PositionWindows(QtWidgets.QWidget, Ui_Form_position):
    def __init__(self, parent=None):
        super(PositionWindows, self).__init__(parent)
        self.setupUi(self)

        axis2_correct = conf.get('Position', 'axis2_correct')
        axis1_up = conf.get('Position', 'axis1_up')
        axis1_down = conf.get('Position', 'axis1_down')
        axis0_simple = conf.get('Position', 'axis0_simple')
        axis0_balance = conf.get('Position', 'axis0_balance')
        axis0_weigh = conf.get('Position', 'axis0_weigh')
        self.lineEdit.setText(axis2_correct)
        self.lineEdit_2.setText(axis1_up)
        self.lineEdit_3.setText(axis1_down)
        self.lineEdit_4.setText(axis0_simple)
        self.lineEdit_5.setText(axis0_balance)
        self.lineEdit_6.setText(axis0_weigh)

        me47_starting_layer = conf.get('Filter_me47', 'starting_layer')
        me47_layer_spacing = conf.get('Filter_me47', 'layer_spacing')
        me47_bit_spacing = conf.get('Filter_me47', 'bit_spacing')
        me47_layer_number = conf.get('Filter_me47', 'layer_number')
        me47_up_distance = conf.get('Filter_me47', 'up_distance')
        self.lineEdit_8.setText(me47_starting_layer)
        self.lineEdit_13.setText(me47_layer_spacing)
        self.lineEdit_14.setText(me47_bit_spacing)
        self.lineEdit_15.setText(me47_layer_number)
        self.lineEdit_19.setText(me47_up_distance)

        ca_starting_layer = conf.get('Filter_ca', 'starting_layer')
        ca_layer_spacing = conf.get('Filter_ca', 'layer_spacing')
        ca_bit_spacing = conf.get('Filter_ca', 'bit_spacing')
        ca_layer_number = conf.get('Filter_ca', 'layer_number')
        ca_up_distance = conf.get('Filter_ca', 'up_distance')
        self.lineEdit_7.setText(ca_starting_layer)
        self.lineEdit_10.setText(ca_layer_spacing)
        self.lineEdit_11.setText(ca_bit_spacing)
        self.lineEdit_12.setText(ca_layer_number)
        self.lineEdit_20.setText(ca_up_distance)

        me90_starting_layer = conf.get('Filter_me90', 'starting_layer')
        me90_layer_spacing = conf.get('Filter_me90', 'layer_spacing')
        me90_bit_spacing = conf.get('Filter_me90', 'bit_spacing')
        me90_layer_number = conf.get('Filter_me90', 'layer_number')
        me90_up_distance = conf.get('Filter_me90', 'up_distance')
        self.lineEdit_9.setText(me90_starting_layer)
        self.lineEdit_17.setText(me90_layer_spacing)
        self.lineEdit_16.setText(me90_bit_spacing)
        self.lineEdit_18.setText(me90_layer_number)
        self.lineEdit_21.setText(me90_up_distance)

        self.pushButton.clicked.connect(self.confirm)
        self.pushButton.setShortcut("enter")  # 绑定快捷键
        self.pushButton_3.clicked.connect(self.confirm2)
        self.pushButton_3.setShortcut("enter")  # 绑定快捷键
        self.pushButton_2.clicked.connect(self.close)
        self.pushButton_2.setShortcut("esc")
        self.pushButton_4.clicked.connect(self.close)
        self.pushButton_4.setShortcut("esc")

    def confirm(self):
        axis2_correct = self.lineEdit.text()
        axis1_up = self.lineEdit_2.text()
        axis1_down = self.lineEdit_3.text()
        axis0_simple = self.lineEdit_4.text()
        axis0_balance = self.lineEdit_5.text()
        axis0_weigh = self.lineEdit_6.text()
        conf.set('Position', 'axis2_correct', axis2_correct)
        conf.set('Position', 'axis1_up', axis1_up)
        conf.set('Position', 'axis1_down', axis1_down)
        conf.set('Position', 'axis0_simple', axis0_simple)
        conf.set('Position', 'axis0_balance', axis0_balance)
        conf.set('Position', 'axis0_weigh', axis0_weigh)

        with open('setting_aws.ini', 'w') as configfile:
            conf.write(configfile)
        self.close()

    def confirm2(self):
        me47_starting_layer = self.lineEdit_8.text()
        conf.set('Filter_me47', 'starting_layer', me47_starting_layer)
        me47_layer_spacing = self.lineEdit_13.text()
        conf.set('Filter_me47', 'layer_spacing', me47_layer_spacing)
        me47_bit_spacing = self.lineEdit_14.text()
        conf.set('Filter_me47', 'bit_spacing', me47_bit_spacing)
        me47_layer_number = self.lineEdit_15.text()
        conf.set('Filter_me47', 'layer_number', me47_layer_number)
        me47_up_distance = self.lineEdit_19.text()
        conf.set('Filter_me47', 'up_distance', me47_up_distance)

        ca_starting_layer = self.lineEdit_7.text()
        conf.set('Filter_ca', 'starting_layer', ca_starting_layer)
        ca_layer_spacing = self.lineEdit_10.text()
        conf.set('Filter_ca', 'layer_spacing', ca_layer_spacing)
        ca_bit_spacing = self.lineEdit_11.text()
        conf.set('Filter_ca', 'bit_spacing', ca_bit_spacing)
        ca_layer_number = self.lineEdit_12.text()
        conf.set('Filter_ca', 'layer_number', ca_layer_number)
        ca_up_distance = self.lineEdit_20.text()
        conf.set('Filter_ca', 'up_distance', ca_up_distance)

        me90_starting_layer = self.lineEdit_9.text()
        conf.set('Filter_me90', 'starting_layer', me90_starting_layer)
        me90_layer_spacing = self.lineEdit_17.text()
        conf.set('Filter_me90', 'layer_spacing', me90_layer_spacing)
        me90_bit_spacing = self.lineEdit_16.text()
        conf.set('Filter_me90', 'bit_spacing', me90_bit_spacing)
        me90_layer_number = self.lineEdit_18.text()
        conf.set('Filter_me90', 'layer_number', me90_layer_number)
        me90_up_distance = self.lineEdit_20.text()
        conf.set('Filter_me90', 'up_distance', me90_up_distance)

        with open('setting_aws.ini', 'w') as configfile:
            conf.write(configfile)
        self.close()


class SetWindows(QtWidgets.QWidget, Ui_Form_system):
    def __init__(self, parent=None):
        super(SetWindows, self).__init__(parent)
        self.setupUi(self)
        time_ion = conf.get('Parameter', 'time_ion')
        time_balance = conf.get('Parameter', 'time_balance')
        weight_interval = conf.get('Parameter', 'weight_interval')
        delay_start = conf.get('Parameter', 'delay_start')
        calibration_value = conf.get('Parameter', 'calibration_value')
        allow_deviation = conf.get('Parameter', 'allow_deviation')
        max_deviation = conf.get('Parameter', 'max_deviation')
        self.spinBox.setValue(int(time_ion))
        self.spinBox_4.setValue(int(time_balance))
        self.spinBox_3.setValue(int(weight_interval))
        self.doubleSpinBox_3.setValue(float(delay_start))
        self.lineEdit.setText(calibration_value)
        self.lineEdit_2.setText(allow_deviation)
        self.lineEdit_3.setText(max_deviation)

        self.pushButton.clicked.connect(self.confirm)
        self.pushButton.setShortcut("enter")  # 绑定快捷键
        self.pushButton_2.clicked.connect(self.close)
        self.pushButton_2.setShortcut("esc")

    def confirm(self):
        tem_set = self.doubleSpinBox_5.value()
        hum_set = self.doubleSpinBox_6.value()
        tem_devia = self.doubleSpinBox.value()
        hum_devia = self.doubleSpinBox_2.value()
        time_ion = self.spinBox.value()
        time_balance = self.spinBox_4.value()
        weight_interval = self.spinBox_3.value()
        delay_start = self.doubleSpinBox_3.value()
        calibration_value = self.lineEdit.text()
        allow_deviation = self.lineEdit_2.text()
        max_deviation = self.lineEdit_3.text()
        if self.checkBox_2.isChecked():
            email_set = self.lineEdit_4.text()
        else:
            email_set = ''
        conf.set('Parameter', 'tem_set', str(tem_set))
        conf.set('Parameter', 'hum_set', str(hum_set))
        conf.set('Parameter', 'tem_devia', str(tem_devia))
        conf.set('Parameter', 'hum_devia', str(hum_devia))
        conf.set('Parameter', 'email', email_set)
        conf.set('Parameter', 'time_ion', str(time_ion))
        conf.set('Parameter', 'time_balance', str(time_balance))
        conf.set('Parameter', 'weight_interval', str(weight_interval))
        conf.set('Parameter', 'delay_start', str(delay_start))
        conf.set('Parameter', 'calibration_value', calibration_value)
        conf.set('Parameter', 'allow_deviation', allow_deviation)
        conf.set('Parameter', 'max_deviation', max_deviation)

        with open('setting_aws.ini', 'w') as configfile:
            conf.write(configfile)
        self.close()


class YingjianWindow(QtWidgets.QWidget, Ui_Form_aws):
    def __init__(self, parent=None):
        super(YingjianWindow, self).__init__(parent)
        self.setupUi(self)
        self.center()  # 界面居中显示
        self.solt_def()

    def center(self):
        screen = QDesktopWidget().screenGeometry()
        size = self.geometry()
        newLeft = (screen.width() - size.width()) / 2
        newTop = ((screen.height() - size.height()) / 2) - 40
        self.move(int(newLeft), int(newTop))

    def solt_def(self):
        self.pushButton_9.clicked.connect(self.axis3_up)
        self.pushButton_10.clicked.connect(self.axis3_down)
        self.pushButton.clicked.connect(self.axis2_move_step)
        self.pushButton_5.clicked.connect(self.axis2_start)
        self.pushButton_6.clicked.connect(self.axis2_stop)
        self.pushButton_17.clicked.connect(self.axis2_reset)
        self.pushButton_2.clicked.connect(self.axis1_position_top)
        self.pushButton_3.clicked.connect(self.axis1_position_up)
        self.pushButton_4.clicked.connect(self.axis1_point_move)
        self.pushButton_13.clicked.connect(self.axis1_up)
        self.pushButton_14.clicked.connect(self.axis1_down)
        self.pushButton_19.clicked.connect(self.axis1_stop)
        self.pushButton_16.clicked.connect(self.axis1_reset)
        self.pushButton_11.clicked.connect(self.axis0_left)
        self.pushButton_12.clicked.connect(self.axis0_right)
        self.pushButton_15.clicked.connect(self.axis0_stop)
        self.pushButton_18.clicked.connect(self.axis0_reset)
        self.pushButton_7.clicked.connect(self.axis0_position_simple)
        self.pushButton_8.clicked.connect(self.axis0_position_balance)
        self.pushButton_20.clicked.connect(self.axis0_position_cover)
        self.pushButton_21.clicked.connect(self.banlance_zero)
        self.pushButton_22.setCheckable(True)
        self.pushButton_22.clicked[bool].connect(self.ion_fan)
        self.pushButton_23.clicked.connect(self.code)

    def axis2_move_step(self):
        self.thread_1 = Thread_1()  # 创建线程
        self.thread_1.start()  # 开始线程

    def axis1_position_top(self):
        self.thread_2 = Thread_2()  # 创建线程
        self.thread_2.start()  # 开始线程

    def axis1_position_up(self):
        self.thread_3 = Thread_3()  # 创建线程
        self.thread_3.start()  # 开始线程

    def axis1_point_move(self):
        self.thread_4 = Thread_4()  # 创建线程
        self.thread_4.start()  # 开始线程

    def axis2_start(self):
        self.thread_5 = Thread_5()  # 创建线程
        self.thread_5.start()  # 开始线程

    def axis2_stop(self):
        self.thread_6 = Thread_6()  # 创建线程
        self.thread_6.start()  # 开始线程

    def axis0_position_simple(self):
        self.thread_7 = Thread_7()  # 创建线程
        self.thread_7.start()  # 开始线程

    def axis0_position_balance(self):
        self.thread_8 = Thread_8()  # 创建线程
        self.thread_8.start()  # 开始线程

    def axis3_up(self):
        self.thread_9 = Thread_9()  # 创建线程
        self.thread_9.start()  # 开始线程

    def axis3_down(self):
        self.thread_10 = Thread_10()  # 创建线程
        self.thread_10.start()  # 开始线程

    def axis0_left(self):
        self.thread_11 = Thread_11()  # 创建线程
        self.thread_11.start()  # 开始线程

    def axis0_right(self):
        self.thread_12 = Thread_12()  # 创建线程
        self.thread_12.start()  # 开始线程

    def axis1_up(self):
        self.thread_13 = Thread_13()  # 创建线程
        self.thread_13.start()  # 开始线程

    def axis1_down(self):
        self.thread_14 = Thread_14()  # 创建线程
        self.thread_14.start()  # 开始线程

    def axis0_stop(self):
        self.thread_15 = Thread_15()  # 创建线程
        self.thread_15.start()  # 开始线程

    def axis1_reset(self):
        self.thread_16 = Thread_16()  # 创建线程
        self.thread_16.start()  # 开始线程

    def axis2_reset(self):
        self.thread_17 = Thread_17()  # 创建线程
        self.thread_17.start()  # 开始线程

    def axis0_reset(self):
        self.thread_18 = Thread_18()  # 创建线程
        self.thread_18.start()  # 开始线程

    def axis1_stop(self):
        self.thread_19 = Thread_19()  # 创建线程
        self.thread_19.start()  # 开始线程

    def axis0_position_cover(self):
        self.thread_20 = Thread_20()  # 创建线程
        self.thread_20.start()  # 开始线程

    def banlance_zero(self):
        self.thread_21 = Thread_21()  # 创建线程
        self.thread_21.start()  # 开始线程

    def ion_fan(self, pressed):
        if pressed:
            self.thread_22_0 = ThreadIonFanOpen()
            self.thread_22_0.start()
        else:
            self.thread_22_1 = ThreadIonFanClose()
            self.thread_22_1.start()

    def code(self):
        self.thread_23 = Thread_23()  # 创建线程
        self.thread_23.start()  # 开始线程


class Thread_1(QThread):  # 2轴移动一位--ok
    def __init__(self):
        super(Thread_1, self).__init__()

    def run(self):
        ip_2048 = "192.168.0.11"
        zmc.connect(ip_2048)  # 连接控制器
        # zmc.read_info()  # 读取控制器信息
        zmc.set_axis_type()
        zmc.set_axis_units()
        zmc.set_op(7, 0)
        zmc.axis_2_move(8572)


class Thread_2(QThread):  # 1轴上升至天平顶--ok
    def __init__(self):
        super(Thread_2, self).__init__()

    def run(self):
        ip_2048 = "192.168.0.11"
        zmc.connect(ip_2048)  # 连接控制器
        # zmc.read_info()  # 读取控制器信息
        zmc.set_axis_type()
        zmc.set_axis_units()
        zmc.set_op(7, 0)
        axis1_up = int(conf.get('Position', 'axis1_up'))
        zmc.axis_1_up(axis1_up)


class Thread_3(QThread):  # 1轴上升至天平底--ok
    def __init__(self):
        super(Thread_3, self).__init__()

    def run(self):
        ip_2048 = "192.168.0.11"
        zmc.connect(ip_2048)  # 连接控制器
        # zmc.read_info()  # 读取控制器信息
        zmc.set_axis_type()
        zmc.set_axis_units()
        zmc.set_op(7, 0)
        axis1_down = int(conf.get('Position', 'axis1_down'))
        zmc.axis_1_up(axis1_down)


class Thread_4(QThread):  # 1轴点动--ok
    def __init__(self):
        super(Thread_4, self).__init__()

    def run(self):
        ip_2048 = "192.168.0.11"
        zmc.connect(ip_2048)  # 连接控制器
        # zmc.read_info()  # 读取控制器信息
        zmc.set_axis_type()
        zmc.set_axis_units()
        zmc.set_op(7, 0)
        zmc.axis_1_move(1000)


class Thread_5(QThread):  # 2轴开始运行--ok
    def __init__(self):
        super(Thread_5, self).__init__()

    def run(self):
        ip_2048 = "192.168.0.11"
        zmc.connect(ip_2048)  # 连接控制器
        # zmc.read_info()  # 读取控制器信息
        zmc.set_axis_type()
        zmc.set_axis_units()
        zmc.set_op(7, 0)
        zmc.set_speed(2, 3000)
        zmc.set_move(2, 1)


class Thread_6(QThread):  # 2轴停止运行--ok
    def __init__(self):
        super(Thread_6, self).__init__()

    def run(self):
        ip_2048 = "192.168.0.11"
        zmc.connect(ip_2048)  # 连接控制器
        # zmc.read_info()  # 读取控制器信息
        zmc.set_axis_type()
        zmc.set_axis_units()
        zmc.set_op(7, 0)
        zmc.set_cancel(2, 0)


class Thread_7(QThread):  # 1轴位置-样品ok
    def __init__(self):
        super(Thread_7, self).__init__()

    def run(self):
        ip_2048 = "192.168.0.11"
        zmc.connect(ip_2048)  # 连接控制器
        # zmc.read_info()  # 读取控制器信息
        zmc.set_axis_type()
        zmc.set_axis_units()
        zmc.set_op(7, 0)
        axis0_simple = int(conf.get('Position', 'axis0_simple'))
        zmc.axis_0_move(axis0_simple)


class Thread_8(QThread):  # ok
    def __init__(self):
        super(Thread_8, self).__init__()

    def run(self):
        ip_2048 = "192.168.0.11"
        zmc.connect(ip_2048)  # 连接控制器
        # zmc.read_info()  # 读取控制器信息
        zmc.set_axis_type()
        zmc.set_axis_units()
        zmc.set_op(7, 0)
        axis0_balance = int(conf.get('Position', 'axis0_balance'))
        zmc.axis_0_move(axis0_balance)


class Thread_9(QThread):  # 天平罩上升--ok
    def __init__(self):
        super(Thread_9, self).__init__()

    def run(self):
        ip_2048 = "192.168.0.11"
        zmc.connect(ip_2048)  # 连接控制器
        # zmc.read_info()  # 读取控制器信息
        zmc.set_axis_type()
        zmc.set_op(7, 0)
        zmc.axis_3_up()


class Thread_10(QThread):  # 天平罩下降--ok
    def __init__(self):
        super(Thread_10, self).__init__()

    def run(self):
        ip_2048 = "192.168.0.11"
        zmc.connect(ip_2048)  # 连接控制器
        # zmc.read_info()  # 读取控制器信息
        zmc.set_axis_type()
        zmc.set_op(7, 0)
        zmc.axis_3_down()


class Thread_11(QThread):  # 机械臂左旋--ok
    def __init__(self):
        super(Thread_11, self).__init__()

    def run(self):
        ip_2048 = "192.168.0.11"
        zmc.connect(ip_2048)  # 连接控制器
        # zmc.read_info()  # 读取控制器信息
        zmc.set_axis_type()
        zmc.set_axis_units()
        zmc.set_op(7, 0)
        zmc.set_move(0, -1)


class Thread_12(QThread):  # 机械臂右旋--ok
    def __init__(self):
        super(Thread_12, self).__init__()

    def run(self):
        ip_2048 = "192.168.0.11"
        zmc.connect(ip_2048)  # 连接控制器
        # zmc.read_info()  # 读取控制器信息
        zmc.set_axis_type()
        zmc.set_axis_units()
        zmc.set_op(7, 0)
        zmc.set_move(0, 1)


class Thread_13(QThread):  # O轴上升
    def __init__(self):
        super(Thread_13, self).__init__()

    def run(self):
        ip_2048 = "192.168.0.11"
        zmc.connect(ip_2048)  # 连接控制器
        # zmc.read_info()  # 读取控制器信息
        zmc.set_axis_type()
        zmc.set_axis_units()
        zmc.set_op(7, 0)
        zmc.set_speed(1, 4000)
        zmc.set_move(1, 1)


class Thread_14(QThread):  # O轴下降
    def __init__(self):
        super(Thread_14, self).__init__()

    def run(self):
        ip_2048 = "192.168.0.11"
        zmc.connect(ip_2048)  # 连接控制器
        # zmc.read_info()  # 读取控制器信息
        zmc.set_axis_type()
        zmc.set_axis_units()
        zmc.set_op(7, 0)
        zmc.set_speed(1, 4000)
        zmc.set_move(1, -1)


class Thread_15(QThread):  # 旋转停止
    def __init__(self):
        super(Thread_15, self).__init__()

    def run(self):
        ip_2048 = "192.168.0.11"
        zmc.connect(ip_2048)  # 连接控制器
        # zmc.read_info()  # 读取控制器信息
        zmc.set_axis_type()
        zmc.set_axis_units()
        zmc.set_op(7, 0)
        zmc.set_cancel(0, 0)


class Thread_16(QThread):  # 上升下降回复
    def __init__(self):
        super(Thread_16, self).__init__()

    def run(self):
        ip_2048 = "192.168.0.11"
        zmc.connect(ip_2048)  # 连接控制器
        # zmc.read_info()  # 读取控制器信息
        zmc.set_axis_type()
        zmc.set_axis_units()
        zmc.set_op(7, 0)
        zmc.axis_1_back()


class Thread_17(QThread):  # 样品架旋转复位--ok
    def __init__(self):
        super(Thread_17, self).__init__()

    def run(self):
        ip_2048 = "192.168.0.11"
        zmc.connect(ip_2048)  # 连接控制器
        # zmc.read_info()  # 读取控制器信息
        zmc.set_axis_type()
        zmc.set_axis_units()
        zmc.set_op(7, 0)
        zmc.axis_2_back()
        axis2_correct = int(conf.get('Position', 'axis2_correct'))
        if axis2_correct == 0:
            pass
        else:
            zmc.axis_2_moveabs(axis2_correct)
            speed = zmc.get_speed(2)
            time.sleep(1 + abs(axis2_correct) / speed)


class Thread_18(QThread):  # 上下运动复位ok
    def __init__(self):
        super(Thread_18, self).__init__()

    def run(self):
        ip_2048 = "192.168.0.11"
        zmc.connect(ip_2048)  # 连接控制器
        # zmc.read_info()  # 读取控制器信息
        zmc.set_axis_type()
        zmc.set_axis_units()
        zmc.set_op(7, 0)
        zmc.set_op(0, 1)
        zmc.set_op(0, 0)
        zmc.set_DPOS(0, 0)


class Thread_19(QThread):  # 上下运动停止ok
    def __init__(self):
        super(Thread_19, self).__init__()

    def run(self):
        ip_2048 = "192.168.0.11"
        zmc.connect(ip_2048)  # 连接控制器
        # zmc.read_info()  # 读取控制器信息
        zmc.set_axis_type()
        zmc.set_axis_units()
        zmc.set_op(7, 0)
        zmc.set_cancel(1, 0)


class Thread_20(QThread):  # 上下运动停止ok
    def __init__(self):
        super(Thread_20, self).__init__()

    def run(self):
        ip_2048 = "192.168.0.11"
        zmc.connect(ip_2048)  # 连接控制器
        # zmc.read_info()  # 读取控制器信息
        zmc.set_axis_type()
        zmc.set_axis_units()
        zmc.set_op(7, 0)
        axis0_weigh = int(conf.get('Position', 'axis0_weigh'))
        zmc.axis_0_move(axis0_weigh)


class Thread_21(QThread):  # 天平置零
    def __init__(self):
        super(Thread_21, self).__init__()

    def run(self):
        zmc.reset_ban()


class Thread_23(QThread):  # 扫码
    def __init__(self):
        super(Thread_23, self).__init__()

    def run(self):
        zmc.code_ban()


class ThreadIonFanOpen(QThread):  # 扫码
    def __init__(self):
        super(ThreadIonFanOpen, self).__init__()

    def run(self):
        zmc.set_op(1, 1)  # 离子风开


class ThreadIonFanClose(QThread):  # 扫码
    def __init__(self):
        super(ThreadIonFanClose, self).__init__()

    def run(self):
        zmc.set_op(1, 0)  # 离子风关


if __name__ == "__main__":
    app = QApplication(sys.argv)
    myWin = MyMainWindow()
    myWin.show()
    sys.exit(app.exec_())
