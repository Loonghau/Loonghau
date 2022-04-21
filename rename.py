# -*- coding: utf-8 -*-
""" 
@Time : 2021/8/6 10:34
@Author : Loonghau XU
@FileName: rename.py
@SoftWare: PyCharm
"""
import os
import time

from openpyxl import load_workbook, Workbook

excelName1 = 'text1.xlsx'
excelName2 = 'result.xlsx'

if os.path.exists(excelName1):
    wb1 = load_workbook(excelName1)
    pass
else:
    wb1 = Workbook()
now_time = time.strftime('%m-%d', time.localtime(time.time()))
if wb1.sheetnames[0] == str(now_time):
    ws2 = wb1[str(now_time)]
    pass
else:
    ws2 = wb1.create_sheet(str(now_time), 0)
    ws2['A1'] = '日期/时间'
    ws2['B1'] = '样品编号'
    ws2['C1'] = '温度'
    ws2['D1'] = '湿度'
    ws2['E1'] = '重量'
    ws2['F1'] = '样品位号'
now_time_1 = time.strftime('%Y-%m-%d %H-%M', time.localtime(time.time()))
ws2.column_dimensions['A'].width = 20
ws2.column_dimensions['B'].width = 15
ws2.column_dimensions['C'].width = 15
ws2.column_dimensions['D'].width = 15
ws2.column_dimensions['E'].width = 15
ws2.column_dimensions['F'].width = 15
list = [str(now_time_1), str(222), str(33), str(4), str(5) + "g", str(6)]
ws2.append(list)
wb1.save(excelName1)
new_name = str(now_time_1) + str(9) + ".xls"
os.rename("text1.xlsx", new_name)

if os.path.exists(excelName2):
    wb2 = load_workbook(excelName2)
    pass
else:
    wb2 = Workbook()
if wb2.sheetnames[0] == str(now_time):
    ws3 = wb2[str(now_time)]
    pass
else:
    ws3 = wb2.create_sheet(str(now_time), 0)
    ws3['A1'] = '日期/时间'
    ws3['B1'] = '样品编号'
    ws3['C1'] = '温度'
    ws3['D1'] = '湿度'
    ws3['E1'] = '重量'
    ws3['F1'] = '样品位号'
now_time_1 = time.strftime('%Y-%m-%d %H-%M', time.localtime(time.time()))
ws3.column_dimensions['A'].width = 20
ws3.column_dimensions['B'].width = 15
ws3.column_dimensions['C'].width = 15
ws3.column_dimensions['D'].width = 15
ws3.column_dimensions['E'].width = 15
ws3.column_dimensions['F'].width = 15
list = [str(now_time_1), str(222), str(33), str(4), str(5) + "g", str(6)]
ws3.append(list)
wb2.save(excelName2)
